"""
app.py — QuantLens · Professional Stock Forecasting Dashboard
=============================================================
Run:  streamlit run app.py
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from typing import Dict, Any, Optional

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG — must be first Streamlit call
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QuantLens · Stock Forecasting",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS — dark fintech aesthetic
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=Space+Grotesk:wght@400;500;600;700&display=swap');

/* Global */
html, body, [class*="css"] {
    font-family: 'Space Grotesk', sans-serif;
    background-color: #0D1117;
    color: #E6EDF3;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #161B22;
    border-right: 1px solid #30363D;
}
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stSlider label,
section[data-testid="stSidebar"] .stTextInput label,
section[data-testid="stSidebar"] .stRadio label {
    color: #8B949E !important;
    font-size: 0.78rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    font-family: 'IBM Plex Mono', monospace;
}

/* Metric cards */
.metric-card {
    background: #161B22;
    border: 1px solid #30363D;
    border-radius: 8px;
    padding: 16px 20px;
    text-align: center;
}
.metric-card .label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.70rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #8B949E;
    margin-bottom: 6px;
}
.metric-card .value {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.60rem;
    font-weight: 600;
    color: #E6EDF3;
    line-height: 1.1;
}
.metric-card .delta {
    font-size: 0.78rem;
    margin-top: 4px;
    font-family: 'IBM Plex Mono', monospace;
}
.delta-up   { color: #3FB950; }
.delta-down { color: #F85149; }
.delta-neu  { color: #8B949E; }

/* Section headers */
.section-header {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.68rem;
    letter-spacing: 0.20em;
    text-transform: uppercase;
    color: #58A6FF;
    border-bottom: 1px solid #30363D;
    padding-bottom: 6px;
    margin: 28px 0 18px 0;
}

/* Result badge */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.72rem;
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    letter-spacing: 0.05em;
}
.badge-green { background: #1a3a22; color: #3FB950; border: 1px solid #3FB950; }
.badge-red   { background: #3a1a1a; color: #F85149; border: 1px solid #F85149; }
.badge-blue  { background: #1a2a3a; color: #58A6FF; border: 1px solid #58A6FF; }
.badge-orange{ background: #3a2a1a; color: #D29922; border: 1px solid #D29922; }

/* Warning / info boxes */
.info-box {
    background: #1a2a3a;
    border-left: 3px solid #58A6FF;
    padding: 10px 16px;
    border-radius: 0 6px 6px 0;
    font-size: 0.84rem;
    margin: 10px 0;
}
.warn-box {
    background: #3a2a1a;
    border-left: 3px solid #D29922;
    padding: 10px 16px;
    border-radius: 0 6px 6px 0;
    font-size: 0.84rem;
    margin: 10px 0;
}

/* Streamlit overrides */
.stButton > button {
    background: #21262D;
    border: 1px solid #30363D;
    color: #E6EDF3;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.82rem;
    border-radius: 6px;
    padding: 8px 24px;
    transition: all 0.15s;
}
.stButton > button:hover {
    background: #58A6FF;
    border-color: #58A6FF;
    color: #0D1117;
}

div[data-testid="stExpander"] {
    border: 1px solid #30363D;
    border-radius: 8px;
    background: #161B22;
}

.stTabs [data-baseweb="tab-list"] {
    background: #161B22;
    border-bottom: 1px solid #30363D;
}
.stTabs [data-baseweb="tab"] {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.80rem;
    letter-spacing: 0.05em;
    color: #8B949E;
}
.stTabs [aria-selected="true"] {
    color: #58A6FF !important;
    border-bottom-color: #58A6FF !important;
}

.stDataFrame { background: #161B22 !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def metric_card(label: str, value: str, delta: str = "", delta_dir: str = "neu") -> str:
    delta_html = f'<div class="delta delta-{delta_dir}">{delta}</div>' if delta else ""
    return f"""
    <div class="metric-card">
        <div class="label">{label}</div>
        <div class="value">{value}</div>
        {delta_html}
    </div>
    """

def badge(text: str, color: str = "blue") -> str:
    return f'<span class="badge badge-{color}">{text}</span>'

def section(title: str):
    st.markdown(f'<div class="section-header">{title}</div>', unsafe_allow_html=True)

def fmt(val, decimals=2, suffix=""):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    return f"{val:.{decimals}f}{suffix}"

def pval_badge(p):
    if p is None or np.isnan(p):
        return badge("N/A", "blue")
    if p < 0.01:
        return badge(f"p={p:.4f} ✗", "red")
    if p < 0.05:
        return badge(f"p={p:.4f} ✗", "orange")
    return badge(f"p={p:.4f} ✓", "green")

try:
    import matplotlib  # noqa: F401
    _HAS_MPL = True
except Exception:
    _HAS_MPL = False

def safe_background_gradient(styler, subset=None, cmap="RdYlGn", vmin=None, vmax=None):
    if not _HAS_MPL:
        return styler
    return styler.background_gradient(subset=subset, cmap=cmap, vmin=vmin, vmax=vmax)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style="padding: 12px 0 20px 0;">
        <div style="font-family:'IBM Plex Mono',monospace; font-size:1.25rem; 
                    font-weight:600; color:#58A6FF; letter-spacing:0.05em;">
            📈 QuantLens
        </div>
        <div style="font-size:0.72rem; color:#8B949E; font-family:'IBM Plex Mono',monospace;
                    letter-spacing:0.10em; margin-top:4px;">
            STOCK FORECASTING SYSTEM
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Data Parameters ──
    st.markdown('<div style="font-family:IBM Plex Mono,monospace;font-size:0.68rem;'
                'letter-spacing:0.15em;text-transform:uppercase;color:#8B949E;'
                'margin-bottom:8px;">Data Parameters</div>', unsafe_allow_html=True)

    symbols_raw = st.text_input(
        "Symbols (comma-separated)",
        value="AAPL,MSFT",
        help="Yahoo Finance tickers: AAPL, TCS.NS, RELIANCE.NS, etc."
    )
    symbols = [s.strip().upper() for s in symbols_raw.split(",") if s.strip()]

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", value=pd.Timestamp("2020-01-01"),
                                   min_value=pd.Timestamp("2000-01-01"))
    with col2:
        end_date = st.date_input("End Date", value=pd.Timestamp("2024-12-31"))

    selected_col = st.selectbox("Price Column", ["Close", "Open", "High", "Low"], index=0)

    st.markdown("---")

    # ── Model Parameters ──
    st.markdown('<div style="font-family:IBM Plex Mono,monospace;font-size:0.68rem;'
                'letter-spacing:0.15em;text-transform:uppercase;color:#8B949E;'
                'margin-bottom:8px;">Model Parameters</div>', unsafe_allow_html=True)

    train_ratio   = st.slider("Train Ratio",   0.50, 0.80, 0.70, 0.05)
    val_ratio     = st.slider("Val Ratio",     0.05, 0.25, 0.15, 0.05)
    future_days   = st.slider("Forecast Horizon (days)", 5, 60, 15, 5)
    refit_every   = st.slider("Refit Every N Steps", 10, 100, 50, 10)
    confidence    = st.slider("Confidence Level", 0.90, 0.99, 0.95, 0.01)
    top_n_compare = st.slider("Models to Plot", 1, 5, 3, 1)

    st.markdown("---")

    # ── Risk Parameters ──
    st.markdown('<div style="font-family:IBM Plex Mono,monospace;font-size:0.68rem;'
                'letter-spacing:0.15em;text-transform:uppercase;color:#8B949E;'
                'margin-bottom:8px;">Risk Parameters</div>', unsafe_allow_html=True)

    var_window = st.slider("VaR Rolling Window (days)", 60, 504, 252, 20)
    var_conf   = st.slider("VaR Confidence Level", 0.90, 0.99, 0.95, 0.01)
    conservative_buffer = st.slider(
        "Conservative Buffer %", 0, 20, 10, 1,
        help="Inflates VaR by this % to reduce breach rate (0 = off)"
    ) / 100.0

    st.markdown("---")
    run_btn = st.button("🚀  Run Analysis", use_container_width=True)

    # ── Runtime status ──
    st.markdown("---")
    try:
        import torch as _torch_chk  # noqa
        _tf_ok = True
    except Exception:
        _tf_ok = False
    try:
        from xgboost import XGBRegressor as _xgb_chk  # noqa
        _xgb_ok = True
    except Exception:
        _xgb_ok = False

    st.markdown('<div style="font-family:IBM Plex Mono,monospace;font-size:0.68rem;'
                'letter-spacing:0.15em;text-transform:uppercase;color:#8B949E;'
                'margin-bottom:8px;">Library Status</div>', unsafe_allow_html=True)
    tf_icon  = "🟢" if _tf_ok  else "🔴"
    xgb_icon = "🟢" if _xgb_ok else "🔴"
    try:
        from core.risk import HAS_ARCH as _has_arch
        _arch_ok = _has_arch
    except Exception:
        _arch_ok = False
    arch_icon = "🟢" if _arch_ok else "🔴"
    try:
        from catboost import CatBoostRegressor as _cat_chk  # noqa
        _cat_ok = True
    except Exception:
        _cat_ok = False
    cat_icon  = "🟢" if _cat_ok else "🔴"
    st.markdown(
        f'<div style="font-family:IBM Plex Mono,monospace;font-size:0.75rem;color:#8B949E;">'
        f'{tf_icon} PyTorch (LSTM/GRU)<br>'
        f'{xgb_icon} XGBoost<br>'
        f'{cat_icon} CatBoost<br>'
        f'{arch_icon} arch (GARCH VaR)</div>',
        unsafe_allow_html=True,
    )
    if not _tf_ok:
        st.caption("Install: `pip install torch`")
    if not _arch_ok:
        st.caption("Install: `pip install arch`")
    if not _cat_ok:
        st.caption("Install: `pip install catboost`")

    st.markdown("---")
    st.markdown("""
    <div style="font-family:IBM Plex Mono,monospace;font-size:0.68rem;
                color:#8B949E;line-height:1.6;">
        <div style="color:#3FB950;font-weight:600;margin-bottom:4px;">
            ✅ Leakage Prevention
        </div>
        Target = price.shift(-1)<br>
        All features backward-looking<br>
        SPY lagged by t-1<br>
        GARCH fitted on train only<br>
        Walk-forward: no future data<br>
        ffill only (no bfill)
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN AREA — Landing
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<div style="display:flex; align-items:baseline; gap:14px; padding:8px 0 4px 0;">
    <span style="font-family:'IBM Plex Mono',monospace; font-size:1.8rem; 
                 font-weight:600; color:#E6EDF3;">QuantLens</span>
    <span style="font-family:'IBM Plex Mono',monospace; font-size:0.75rem; 
                 color:#8B949E; letter-spacing:0.15em; text-transform:uppercase;">
        Multi-Model Stock Forecasting &amp; Risk Analytics
    </span>
</div>
<div style="height:2px; background:linear-gradient(90deg,#58A6FF,#3FB950,transparent); 
            margin-bottom:24px;"></div>
""", unsafe_allow_html=True)

if not run_btn:
    st.markdown("""
    <div style="background:#161B22; border:1px solid #30363D; border-radius:10px;
                padding:32px; max-width:680px; margin:40px auto; text-align:center;">
        <div style="font-size:2.5rem; margin-bottom:12px;">📊</div>
        <div style="font-family:'IBM Plex Mono',monospace; font-size:1.05rem; 
                    font-weight:500; color:#58A6FF; margin-bottom:8px;">
            Professional-grade forecasting at your fingertips
        </div>
        <div style="color:#8B949E; font-size:0.88rem; line-height:1.7;">
            Configure your symbols and parameters in the sidebar, then hit 
            <strong style="color:#3FB950;">Run Analysis</strong> to train 
            ARIMA, Log-ARIMA, Holt-Winters, XGBoost, GBM, RandomForest, 
            LSTM &amp; GRU models — with full walk-forward validation, 
            VaR backtesting, and prediction bands.
        </div>
        <div style="margin-top:20px; display:flex; justify-content:center; gap:8px; flex-wrap:wrap;">
            <span style="background:#1a2a3a;border:1px solid #58A6FF;border-radius:20px;
                         padding:4px 12px;font-size:0.72rem;font-family:IBM Plex Mono,monospace;
                         color:#58A6FF;">Walk-Forward CV</span>
            <span style="background:#1a3a22;border:1px solid #3FB950;border-radius:20px;
                         padding:4px 12px;font-size:0.72rem;font-family:IBM Plex Mono,monospace;
                         color:#3FB950;">Kupiec + Christoffersen</span>
            <span style="background:#3a2a1a;border:1px solid #D29922;border-radius:20px;
                         padding:4px 12px;font-size:0.72rem;font-family:IBM Plex Mono,monospace;
                         color:#D29922;">Prediction Bands</span>
            <span style="background:#2a1a3a;border:1px solid #BC8CFF;border-radius:20px;
                         padding:4px 12px;font-size:0.72rem;font-family:IBM Plex Mono,monospace;
                         color:#BC8CFF;">Directional Accuracy</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE EXECUTION
# ─────────────────────────────────────────────────────────────────────────────

from core.data   import load_data, load_market_index, prepare_splits, prepare_order_info, run_walk_forward_cv
from core.models import (train_all_models, evaluate_on_test, compute_metrics,
                         bootstrap_prediction_bands, skill_score,
                         widen_bands_for_horizon, compute_signals,
                         compute_strategy_backtest, extract_garch_vol_oos)
from core.risk   import (full_var_backtest, var_historical, var_parametric,
                         var_ewma, var_garch, var_cornish_fisher, conservative_var,
                         compute_returns, expected_shortfall, HAS_ARCH,
                         run_all_var_methods_parallel, select_best_var_method,
                         METHOD_LABELS)
from core.charts import (
    chart_price, chart_forecast, chart_model_comparison,
    chart_var_breaches, chart_var_all_methods, chart_var_method_comparison,
    chart_walk_forward, chart_directional_accuracy,
    chart_future_forecast, chart_returns_dist, chart_feature_importance, chart_residuals,
    chart_rolling_sharpe,
    THEME
)

# ── Step 1: Load data ──
with st.spinner("⬇️  Downloading market data…"):
    data = load_data(symbols, str(start_date), str(end_date))
    # S&P 500 as market index — used as lagged feature (no leakage)
    market_index = load_market_index(str(start_date), str(end_date))
    if market_index is not None:
        st.caption("📊 S&P 500 market index loaded for feature engineering.")

if not data:
    st.error("No data loaded. Please check your symbols and date range.")
    st.stop()

loaded_syms = list(data.keys())
st.success(f"Loaded: {', '.join(loaded_syms)}")

# ── Step 2: Prepare splits & order info ──
with st.spinner("🔬  Stationarity analysis & ARIMA order selection…"):
    splits     = prepare_splits(data, selected_col, train_ratio, val_ratio)
    order_info = prepare_order_info(splits, selected_col)

# ── Step 3: Train models ──
all_results: Dict[str, Any] = {}
for sym in loaded_syms:
    progress_bar = st.progress(0, text=f"Training models for {sym}…")
    status_msgs = []

    def update_progress(msg):
        status_msgs.append(msg)
        pct = min(len(status_msgs) / 8, 1.0)
        progress_bar.progress(pct, text=f"{sym}: {msg}")

    sp = splits[sym]
    all_results[sym] = train_all_models(
        sym, sp["train"], sp["val"], sp["test"],
        data[sym], order_info, selected_col,
        refit_every=refit_every, confidence=confidence,
        progress_callback=update_progress,
        market_ret=market_index,
    )
    progress_bar.empty()
    # Surface DL error if any
    dl_err = all_results[sym].pop("__dl_error__", None)
    if dl_err:
        st.warning(f"⚠️ {sym} DL models: {dl_err}")

# ── Step 4: Test evaluation ──
test_results: Dict[str, Any] = {}
final_winners: Dict[str, str] = {}

with st.spinner("📊  Evaluating top models on held-out test set…"):
    for sym in loaded_syms:
        sp    = splits[sym]
        res   = all_results[sym]

        sorted_val = sorted(
            [(n, r) for n, r in res.items()
             if not n.startswith("__") and not np.isnan(r.get("mape", np.nan))],
            key=lambda x: x[1]["mape"]
        )
        # Include hybrid and direction models even if not top-5 by MAPE
        hybrid_names = {"XGB+GARCH", "Direction+Vol", "Direction-Vote",
                        "XGB-Returns", "RF-Returns", "CatBoost-Returns"}
        top5 = sorted_val[:5]
        top5_names = {n for n, _ in top5}
        for nm in hybrid_names:
            if nm in res and nm not in top5_names:
                top5.append((nm, res[nm]))

        # Compute OOS GARCH vol for test set (leakage-free)
        garch_vol_test = None
        try:
            garch_vol_test = extract_garch_vol_oos(
                pd.concat([sp["train"], sp["val"]]), sp["test"])
        except Exception:
            pass

        test_results[sym] = evaluate_on_test(
            sym, top5, sp["train"], sp["val"], sp["test"],
            data[sym], order_info, all_results, selected_col,
            refit_every=refit_every, confidence=confidence,
            market_ret=market_index,
            garch_vol_test=garch_vol_test,
        )

        if test_results[sym]:
            best = min(test_results[sym], key=lambda k: test_results[sym][k]["mape"])
            final_winners[sym] = best

# ── Step 5: Walk-Forward CV ──
wf_results: Dict[str, Any] = {}
with st.spinner("🔄  Walk-forward cross-validation…"):
    for sym in loaded_syms:
        price = splits[sym]["full"]
        wf_results[sym] = run_walk_forward_cv(
            sym, price, order_info, n_splits=5,
            refit_every=refit_every, confidence=confidence,
        )

# ── Step 6: VaR Backtest — all methods in parallel ──
var_results:      Dict[str, Any] = {}   # {sym -> {method -> result}}
var_best_method:  Dict[str, str] = {}   # {sym -> best_method_name}

with st.spinner("🛡️  Running all VaR methods in parallel…"):
    for sym in loaded_syms:
        price = splits[sym]["full"]
        all_vr = run_all_var_methods_parallel(
            price,
            confidence=var_conf,
            var_window=var_window,
            conservative_buffer=conservative_buffer,
        )
        best = select_best_var_method(all_vr)
        var_results[sym]     = all_vr
        var_best_method[sym] = best

# ── Step 7: Future forecasts ──
future_forecasts: Dict[str, Any] = {}
with st.spinner("🔮  Generating future forecasts…"):
    for sym in loaded_syms:
        if sym not in final_winners:
            continue
        sp         = splits[sym]
        full_s     = sp["full"]
        best_name  = final_winners[sym]

        # Use val result info (train on full, forecast forward)
        best_info  = all_results[sym].get(best_name, {})
        model_type = best_info.get("type", "arima")

        try:
            from statsmodels.tsa.statespace.sarimax import SARIMAX
            from statsmodels.tsa.holtwinters import ExponentialSmoothing
            from core.models import SARIMAXWrapper, create_sequences
            import numpy as np

            if model_type in ("arima", "log_arima"):
                order = best_info.get("order", order_info[sym]["order"])
                if model_type == "log_arima":
                    m = SARIMAX(np.log(full_s), order=order, seasonal_order=(0,0,0,0),
                                enforce_stationarity=False, enforce_invertibility=False)
                    fitted = m.fit(disp=False, maxiter=500)
                    fc = fitted.get_forecast(steps=future_days)
                    future_pred  = np.exp(fc.predicted_mean)
                    future_lower = np.exp(fc.conf_int(alpha=1-confidence).iloc[:, 0])
                    future_upper = np.exp(fc.conf_int(alpha=1-confidence).iloc[:, 1])
                else:
                    m = SARIMAX(full_s, order=order, seasonal_order=(0,0,0,0),
                                enforce_stationarity=False, enforce_invertibility=False)
                    fitted = m.fit(disp=False, maxiter=500)
                    fc = fitted.get_forecast(steps=future_days)
                    future_pred  = fc.predicted_mean
                    ci = fc.conf_int(alpha=1-confidence)
                    future_lower = ci.iloc[:, 0]
                    future_upper = ci.iloc[:, 1]

            elif model_type == "ml":
                from core.models import compute_features
                model_obj    = best_info["model_obj"]
                feature_cols = best_info.get("feature_cols")
                feat_df      = compute_features(data[sym], selected_col, n_lags=15)
                if feature_cols is None:
                    feature_cols = [c for c in feat_df.columns if c != "target"]
                last_row = feat_df.iloc[-1:].copy()
                preds = []
                for i in range(future_days):
                    x_row = last_row[feature_cols].replace([np.inf,-np.inf], np.nan).ffill().bfill()
                    pred = model_obj.predict(x_row)[0]
                    preds.append(pred)
                    new_row = last_row.copy()
                    for lag in range(15, 1, -1):
                        if f"lag_{lag}" in new_row.columns:
                            new_row[f"lag_{lag}"] = new_row.get(f"lag_{lag-1}", new_row[f"lag_{lag}"])
                    new_row["lag_1"] = pred
                    last_row = new_row
                future_pred  = pd.Series(preds, dtype=float)
                # Bootstrap bands on test residuals
                if sym in test_results and best_name in test_results[sym]:
                    tr = test_results[sym][best_name]
                    residuals = sp["test"].loc[tr["pred"].index] - tr["pred"]
                    alpha = 1 - confidence
                    q_lo, q_hi = residuals.quantile(alpha/2), residuals.quantile(1-alpha/2)
                    future_lower = future_pred + q_lo
                    future_upper = future_pred + q_hi
                else:
                    future_lower = future_pred * 0.97
                    future_upper = future_pred * 1.03

            elif model_type in ("lstm", "gru"):
                model_obj  = best_info["model_obj"]
                scaler     = best_info["scaler"]
                seq_length = best_info["seq_length"]
                scaled     = scaler.transform(full_s.values.reshape(-1, 1))
                last_seq   = scaled[-seq_length:].copy()
                preds = []
                for i in range(future_days):
                    x = last_seq.reshape(1, seq_length, 1)
                    p_scaled = model_obj.predict(x, verbose=0)[0][0]
                    preds.append(p_scaled)
                    last_seq = np.append(last_seq[1:], [[p_scaled]], axis=0)
                future_pred  = pd.Series(
                    scaler.inverse_transform(np.array(preds).reshape(-1, 1)).flatten(),
                    dtype=float
                )
                future_lower = future_pred * 0.97
                future_upper = future_pred * 1.03

            else:  # holt_winters / ensemble fallback
                hw = ExponentialSmoothing(full_s, trend="add", initialization_method="estimated")
                hw_fitted = hw.fit(optimized=True)
                future_pred  = hw_fitted.forecast(future_days)
                future_lower = future_pred * 0.97
                future_upper = future_pred * 1.03

            last_date    = full_s.index[-1]
            future_dates = pd.bdate_range(start=last_date + pd.Timedelta(days=1), periods=future_days)
            future_pred  = pd.Series(future_pred.values, index=future_dates)
            future_lower = pd.Series(
                future_lower.values if hasattr(future_lower, "values") else future_lower,
                index=future_dates
            )
            future_upper = pd.Series(
                future_upper.values if hasattr(future_upper, "values") else future_upper,
                index=future_dates
            )

            future_forecasts[sym] = {
                "pred":  future_pred, "lower": future_lower, "upper": future_upper,
                "model": best_name,  "type":  model_type,
            }
        except Exception as e:
            st.warning(f"{sym} future forecast failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# SYMBOL TABS
# ─────────────────────────────────────────────────────────────────────────────

sym_tabs = st.tabs([f"📈 {sym}" for sym in loaded_syms])

for sym_tab, sym in zip(sym_tabs, loaded_syms):
    with sym_tab:
        sp         = splits[sym]
        res        = all_results[sym]
        t_res      = test_results.get(sym, {})
        winner     = final_winners.get(sym)
        w_res      = wf_results.get(sym, [])
        all_vr     = var_results.get(sym, {})
        best_vm    = var_best_method.get(sym, "historical")
        vr         = all_vr.get(best_vm, {})   # best method result for summary cards
        ff         = future_forecasts.get(sym, {})

        # ── SUMMARY CARDS ──────────────────────────────────────────────────
        section("Summary")
        best_test = t_res.get(winner, {}) if winner else {}

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        cols = [c1, c2, c3, c4, c5, c6]

        metrics_html = [
            metric_card("Best Model", winner.split("(")[0][:12] if winner else "—",
                        badge(res.get(winner, {}).get("type", "").upper(), "blue") if winner else "",
                        "neu"),
            metric_card("Test MAPE",
                        fmt(best_test.get("mape"), 2, "%"),
                        "↓ lower is better", "neu"),
            metric_card("Test RMSE",
                        fmt(best_test.get("rmse"), 2),
                        "", "neu"),
            metric_card("Dir. Accuracy",
                        fmt(best_test.get("da"), 1, "%"),
                        "↑ above 50% = edge",
                        "up" if (best_test.get("da") or 0) > 55 else
                        "down" if (best_test.get("da") or 0) < 50 else "neu"),
            metric_card("VaR Breaches",
                        f"{vr.get('n_breaches', 'N/A')} / {vr.get('n_obs', 'N/A')}",
                        f"exp: {vr.get('exp_rate', 0)*100:.1f}%", "neu"),
            metric_card("Forecast",
                        fmt(ff["pred"].iloc[-1] if ff else None, 2),
                        f"+{future_days}d outlook", "neu"),
        ]
        for col, html in zip(cols, metrics_html):
            col.markdown(html, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── MAIN TABS ──────────────────────────────────────────────────────
        main_tabs = st.tabs([
            "🕯️ Price", "🔮 Forecast", "📊 Model Comparison",
            "🔄 Walk-Forward CV", "🛡️ Risk & VaR",
            "📉 Residuals", "🌅 Future Forecast",
            "📡 Signals", "⚡ Strategy Backtest"
        ])

        # ── Tab 0: Price ──
        with main_tabs[0]:
            section("Price History — Train / Val / Test Split")
            st.plotly_chart(
                chart_price(sp["train"], sp["val"], sp["test"], sym, selected_col),
                use_container_width=True,
            )

            col_l, col_r = st.columns(2)
            with col_l:
                section("ARIMA Order Selection")
                oi = order_info[sym]
                st.markdown(f"""
                <div class="info-box">
                    <strong>ADF p-value:</strong> {oi['stat_info']['adf_p']:.4f} &nbsp;|&nbsp;
                    <strong>KPSS p-value:</strong> {oi['stat_info']['kpss_p']:.4f}<br>
                    <strong>Stationarity:</strong> {'✅ Stationary' if oi['is_stationary'] else '⚠️ Non-Stationary (d=1)'}
                    &nbsp;|&nbsp; <strong>Best ARIMA Order:</strong> {oi['order']}
                </div>
                """, unsafe_allow_html=True)
            with col_r:
                section("Dataset Info")
                st.markdown(f"""
                <div class="info-box">
                    <strong>Symbol:</strong> {sym} &nbsp;|&nbsp;
                    <strong>Column:</strong> {selected_col}<br>
                    <strong>Train:</strong> {len(sp['train'])} bars &nbsp;|&nbsp;
                    <strong>Val:</strong> {len(sp['val'])} bars &nbsp;|&nbsp;
                    <strong>Test:</strong> {len(sp['test'])} bars<br>
                    <strong>Date range:</strong> {sp['full'].index[0].date()} → {sp['full'].index[-1].date()}
                </div>
                """, unsafe_allow_html=True)

        # ── Tab 1: Forecast ──
        with main_tabs[1]:
            section("Validation Set — Actual vs Predicted (Top Models with Prediction Bands)")
            show_bands = st.toggle("Show Prediction Bands", value=True, key=f"bands_{sym}")
            st.plotly_chart(
                chart_forecast(sp["val"], res, sym, top_n=top_n_compare,
                               show_bands=show_bands, history=sp["train"]),
                use_container_width=True,
            )

            if t_res:
                section("Test Set — Actual vs Predicted")
                st.plotly_chart(
                    chart_forecast(sp["test"], t_res, sym, top_n=top_n_compare,
                                   show_bands=show_bands, history=pd.concat([sp["train"], sp["val"]])),
                    use_container_width=True,
                )

                section("Directional Accuracy")
                st.plotly_chart(
                    chart_directional_accuracy({**res, **t_res}, sym),
                    use_container_width=True,
                )
                st.markdown("""
                <div class="info-box">
                    <strong>Directional Accuracy (DA)</strong> measures the % of periods where the model 
                    correctly predicts the direction of price movement (up or down). 
                    Random guessing = 50%. DA &gt; 55% is generally considered meaningful 
                    for equity forecasting. Note: high DA with high MAPE can indicate the model 
                    captures trend but not magnitude.
                </div>
                """, unsafe_allow_html=True)

        # ── Tab 2: Model Comparison ──
        with main_tabs[2]:
            section("Validation Set — All Model Performance")
            st.plotly_chart(chart_model_comparison(res, sym), use_container_width=True)

            section("Validation Metrics Table")
            rows = []
            naive_mape = None
            if "Naive-LastValue" in res and isinstance(res["Naive-LastValue"], dict):
                naive_mape = res["Naive-LastValue"].get("mape", np.nan)

            for name, r in res.items():
                if name.startswith("__") or not isinstance(r, dict):
                    continue
                if r.get("mape") is None:
                    continue
                ss = (skill_score(r.get("mape", np.nan), naive_mape)
                      if naive_mape and not np.isnan(naive_mape) else np.nan)
                rows.append({
                    "Model": name, "Type": r.get("type",""),
                    "Val MAPE %": round(r.get("mape", np.nan), 3),
                    "Val RMSE":   round(r.get("rmse", np.nan), 3),
                    "Val MAE":    round(r.get("mae", np.nan), 3),
                    "DA %":       round(r.get("da", np.nan), 1),
                    "Skill Score": round(ss, 3) if not np.isnan(ss) else "—",
                    "Beats Naive": "✅" if (not np.isnan(ss) and ss > 0) else "❌",
                    "Train MAPE %": round(r.get("train_mape", np.nan), 3),
                    "Overfit Score": round(r.get("overfit_score", np.nan), 3),
                })
            if rows:
                df_table = pd.DataFrame(rows).sort_values("Val MAPE %").reset_index(drop=True)
                styler = df_table.style
                styler = safe_background_gradient(styler, subset=["Val MAPE %"], cmap="RdYlGn_r")
                styler = safe_background_gradient(styler, subset=["DA %"], cmap="RdYlGn")
                st.dataframe(
                    styler.format(na_rep="N/A"),
                    use_container_width=True,
                )
                # Skill score summary
                n_beat = df_table["Beats Naive"].eq("✅").sum()
                total  = len(df_table)
                st.markdown(
                    f'<div class="info-box">📊 <strong>{n_beat}/{total}</strong> models beat '
                    f'the Naive-LastValue baseline. Skill Score &gt; 0 = genuine predictive value.</div>',
                    unsafe_allow_html=True,
                )
                st.download_button(
                    label="⬇️ Download Validation Metrics CSV",
                    data=df_table.to_csv(index=False),
                    file_name=f"{sym}_validation_metrics.csv",
                    mime="text/csv",
                    key=f"dl_val_metrics_{sym}",
                )

            if t_res:
                section("Test Set — Top Model Performance")
                rows_t = []
                for name, r in t_res.items():
                    rows_t.append({
                        "Model": name, "Type": r.get("type",""),
                        "Test MAPE %": round(r.get("mape", np.nan), 3),
                        "Test RMSE":   round(r.get("rmse", np.nan), 3),
                        "Test MAE":    round(r.get("mae", np.nan), 3),
                        "DA %":        round(r.get("da", np.nan), 1),
                    })
                if rows_t:
                    df_t = pd.DataFrame(rows_t).sort_values("Test MAPE %").reset_index(drop=True)
                    styler_t = df_t.style
                    styler_t = safe_background_gradient(styler_t, subset=["Test MAPE %"], cmap="RdYlGn_r")
                    styler_t = safe_background_gradient(styler_t, subset=["DA %"], cmap="RdYlGn")
                    st.dataframe(
                        styler_t.format(na_rep="N/A"),
                        use_container_width=True,
                    )
                    st.download_button(
                        label="⬇️ Download Test Metrics CSV",
                        data=df_t.to_csv(index=False),
                        file_name=f"{sym}_test_metrics.csv",
                        mime="text/csv",
                        key=f"dl_test_metrics_{sym}",
                    )

            # Feature importance for best ML model
            ml_results = {n: r for n, r in res.items()
                          if isinstance(r, dict) and r.get("type") == "ml"
                          and "feature_importance" in r}
            if ml_results:
                best_ml = min(ml_results, key=lambda k: ml_results[k]["mape"])
                section(f"Feature Importance — {best_ml}")
                st.plotly_chart(
                    chart_feature_importance(ml_results[best_ml]["feature_importance"],
                                             best_ml, sym),
                    use_container_width=True,
                )

            # ── Benchmark comparison ──────────────────────────────────────────
            section("Benchmark Comparison — Does Your Model Have Skill?")
            st.markdown("""
            <div class="info-box">
                <strong>Naive baselines</strong> are the minimum bar every model must clear.
                A model that cannot beat a simple "use yesterday's price" is not adding value.<br>
                <strong>Skill Score</strong> = 1 − (model MAPE / naive MAPE).
                &gt; 0 = beats baseline · = 0 = tied · &lt; 0 = worse than naive.
            </div>
            """, unsafe_allow_html=True)

            naive_names = ["Naive-LastValue", "Naive-Drift", "Naive-Mean"]
            naive_rows  = []
            for nm in naive_names:
                nr = res.get(nm, {})
                if not isinstance(nr, dict) or not nr.get("mape"):
                    continue
                naive_rows.append({
                    "Baseline": nm,
                    "MAPE %": round(nr.get("mape", np.nan), 3),
                    "RMSE":   round(nr.get("rmse", np.nan), 3),
                    "DA %":   round(nr.get("da",   np.nan), 1),
                })
            if naive_rows:
                cols_b = st.columns(len(naive_rows))
                for col_b, nr in zip(cols_b, naive_rows):
                    col_b.markdown(
                        metric_card(nr["Baseline"].replace("Naive-",""),
                                    f"MAPE {nr['MAPE %']}%",
                                    f"DA {nr['DA %']}%", "neu"),
                        unsafe_allow_html=True,
                    )

            # Show how many non-naive models beat Naive-LastValue
            baseline_mape = res.get("Naive-LastValue", {}).get("mape", np.nan)
            if not np.isnan(baseline_mape if baseline_mape else np.nan):
                beat_count = sum(
                    1 for n, r in res.items()
                    if isinstance(r, dict) and r.get("type") not in ("naive", None)
                    and not n.startswith("__")
                    and not np.isnan(r.get("mape", np.nan))
                    and r["mape"] < baseline_mape
                )
                total_non_naive = sum(
                    1 for n, r in res.items()
                    if isinstance(r, dict) and r.get("type") not in ("naive", None)
                    and not n.startswith("__")
                    and not np.isnan(r.get("mape", np.nan))
                )
                color = "green" if beat_count > total_non_naive // 2 else "orange"
                st.markdown(
                    f'<div class="{"info-box" if color == "green" else "warn-box"}">'
                    f'<strong>{beat_count} / {total_non_naive}</strong> models beat '
                    f'Naive-LastValue (MAPE {baseline_mape:.3f}%). '
                    f'{"✅ Most models show genuine predictive skill." if beat_count > total_non_naive // 2 else "⚠️ Fewer than half the models beat naive — consider the market regime."}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Tab 3: Walk-Forward CV ──
        with main_tabs[3]:
            section("Walk-Forward Cross-Validation (Expanding Window, 5 Folds)")
            st.markdown("""
            <div class="info-box">
                Expanding-window walk-forward CV trains on all data up to fold start, 
                then tests on the next window — no data leakage. Reference model: AutoARIMA.
                This tests generalization across different market regimes.
            </div>
            """, unsafe_allow_html=True)

            if w_res:
                st.plotly_chart(chart_walk_forward(w_res, sym), use_container_width=True)

                df_wf = pd.DataFrame(w_res)
                col_l, col_r = st.columns([2, 1])
                with col_l:
                    section("Fold Details")
                    st.dataframe(
                        df_wf.style.format({"mape":"{:.3f}","rmse":"{:.3f}",
                                            "mae":"{:.3f}","da":"{:.1f}"}),
                        use_container_width=True,
                    )
                with col_r:
                    section("CV Summary")
                    for metric in ["mape", "rmse", "mae", "da"]:
                        if metric in df_wf.columns:
                            mu = df_wf[metric].mean()
                            sd = df_wf[metric].std()
                            st.markdown(
                                metric_card(metric.upper(), f"{mu:.3f}",
                                            f"± {sd:.3f} std", "neu"),
                                unsafe_allow_html=True,
                            )
                            st.markdown("<br>", unsafe_allow_html=True)
            else:
                st.warning("Walk-forward CV did not produce results (need more data).")

        # ── Tab 4: Risk & VaR ──
        with main_tabs[4]:
            best_label = METHOD_LABELS.get(best_vm, best_vm)
            section(f"🛡️ VaR Analysis — All Methods · Auto-selected: {best_label}")

            st.markdown(f"""
            <div class="info-box">
                All 7 VaR methods were evaluated <strong>in parallel</strong>:
                Historical, Parametric, EWMA, GARCH(1,1)-t, GJR-GARCH(1,1,1),
                Cornish-Fisher, and Regime-Conditional.<br>
                The best method is selected automatically by a composite score:
                Kupiec p-value (40%) + breach rate proximity (40%) + Christoffersen p-value (20%).<br>
                <strong>Auto-selected: {best_label}</strong>
                &nbsp;·&nbsp; Score: {all_vr.get(best_vm, {}).get('score', 0):.3f}
                {f"&nbsp;·&nbsp; Conservative buffer: +{conservative_buffer*100:.0f}%"
                 if conservative_buffer > 0 else ""}
            </div>
            """, unsafe_allow_html=True)

            if not all_vr or all(("error" in r) for r in all_vr.values()):
                st.error("All VaR methods failed. Try a shorter window or more data.")
            else:
                # ── Method comparison bar chart ───────────────────────────
                section("Method Comparison — All 7 Methods")
                st.plotly_chart(
                    chart_var_method_comparison(all_vr, sym, var_conf, best_vm),
                    use_container_width=True,
                )

                # ── Overlay chart ─────────────────────────────────────────
                section("VaR Overlay — All Methods on One Chart")
                best_ret = all_vr.get(best_vm, {}).get("returns", pd.Series(dtype=float))
                st.plotly_chart(
                    chart_var_all_methods(best_ret, all_vr, sym, var_conf, best_vm),
                    use_container_width=True,
                )

                # ── Individual breach chart for best method ───────────────
                section(f"Breach Detail — {best_label} (Auto-selected Best)")
                if "error" not in vr and "returns" in vr:
                    st.plotly_chart(
                        chart_var_breaches(
                            vr["returns"], vr["var_series"],
                            vr["breaches"], sym, var_conf, best_vm),
                        use_container_width=True,
                    )

                # ── Returns distribution ──────────────────────────────────
                if "returns" in vr:
                    ret = vr["returns"]
                    var95_val = float(vr["var_series"].mean()) if not vr["var_series"].isna().all() else 0
                    var99_series = var_historical(ret, 0.99, window=var_window).dropna()
                    var99_val = float(var99_series.mean()) if len(var99_series) else 0
                    st.plotly_chart(
                        chart_returns_dist(ret, sym, var95_val, var99_val),
                        use_container_width=True,
                    )

                # ── Full results table ────────────────────────────────────
                section("All Methods — Detailed Results")
                table_rows = []
                for method, result in all_vr.items():
                    if "error" in result:
                        table_rows.append({
                            "Method": METHOD_LABELS.get(method, method),
                            "Score": "—", "Obs Rate %": "—", "Exp Rate %": "—",
                            "N Breaches": "—", "Kupiec p": "—",
                            "Kupiec Pass": "—", "CC p": "—", "CC Pass": "—",
                            "ES (CVaR %)": "—", "Best": "",
                        })
                        continue
                    kup = result.get("kupiec", {})
                    cc  = result.get("christoffersen", {})
                    table_rows.append({
                        "Method": METHOD_LABELS.get(method, method),
                        "Score": round(result.get("score", 0), 3),
                        "Obs Rate %": round(result.get("obs_rate", 0)*100, 3),
                        "Exp Rate %": round(result.get("exp_rate", 0)*100, 3),
                        "N Breaches": f"{result.get('n_breaches','?')} / {result.get('n_obs','?')}",
                        "Kupiec p": round(kup.get("p_value", 0), 4),
                        "Kupiec Pass": "✅" if not kup.get("reject_h0", True) else "❌",
                        "CC p": round(cc.get("p_value_cc", 0), 4),
                        "CC Pass": "✅" if not cc.get("reject_h0_cc", True) else "❌",
                        "ES (CVaR %)": round(float(result.get("es", 0))*100, 4),
                        "Best": "★" if method == best_vm else "",
                    })

                df_methods = pd.DataFrame(table_rows).sort_values(
                    "Score", ascending=False).reset_index(drop=True)
                st.dataframe(df_methods, use_container_width=True, hide_index=True)

                # ── Stat test detail for best method ─────────────────────
                section(f"Statistical Tests — {best_label}")
                kup  = vr.get("kupiec", {})
                chri = vr.get("christoffersen", {})
                buc  = vr.get("binomial_uc", {})

                st.markdown(f"""
                <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;">
                    <div class="metric-card">
                        <div class="label">Kupiec POF Test</div>
                        <div style="margin:8px 0;">{pval_badge(kup.get('p_value'))}</div>
                        <div style="font-size:0.78rem;color:#8B949E;font-family:IBM Plex Mono,monospace;">
                            LR stat: {fmt(kup.get('lr_stat'), 3)}<br>
                            Obs: {fmt(kup.get('obs_rate',0)*100, 2)}%
                            (exp: {fmt(kup.get('exp_rate',0)*100, 2)}%)<br>
                            Breaches: {kup.get('n_breaches','N/A')} / {kup.get('n','N/A')}
                        </div>
                    </div>
                    <div class="metric-card">
                        <div class="label">Christoffersen CC</div>
                        <div style="margin:8px 0;">{pval_badge(chri.get('p_value_cc'))}</div>
                        <div style="font-size:0.78rem;color:#8B949E;font-family:IBM Plex Mono,monospace;">
                            LR_cc: {fmt(chri.get('lr_cc'), 3)}<br>
                            LR_ind: {fmt(chri.get('lr_ind'), 3)}<br>
                            π₀₁: {fmt(chri.get('pi01'), 3)} · π₁₁: {fmt(chri.get('pi11'), 3)}
                        </div>
                    </div>
                    <div class="metric-card">
                        <div class="label">Binomial UC Test</div>
                        <div style="margin:8px 0;">{pval_badge(buc.get('p_value'))}</div>
                        <div style="font-size:0.78rem;color:#8B949E;font-family:IBM Plex Mono,monospace;">
                            Obs: {fmt(buc.get('obs_rate',0)*100, 2)}%
                            (exp: {fmt(buc.get('exp_rate',0)*100, 2)}%)<br>
                            n = {buc.get('n','N/A')} · k = {buc.get('n_breaches','N/A')}
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                es = vr.get("es", np.nan)
                col_l, col_r = st.columns(2)
                with col_l:
                    section("Expected Shortfall (CVaR)")
                    st.markdown(f"""
                    <div class="info-box">
                        <strong>ES ({var_conf*100:.0f}%):</strong>
                        {fmt(es*100, 3)}% average loss beyond VaR threshold.<br>
                        ES is a coherent risk measure and captures tail risk beyond VaR.
                    </div>
                    """, unsafe_allow_html=True)
                with col_r:
                    section("Scoring Methodology")
                    st.markdown("""
                    <div class="info-box">
                        <strong>Composite Score</strong> = 0.40 × Kupiec p
                        + 0.40 × breach proximity + 0.20 × Christoffersen p<br>
                        <strong>✅</strong> p ≥ 0.05 → model not rejected<br>
                        <strong>❌</strong> p &lt; 0.05 → VaR mis-specified<br>
                        Score closest to 1.0 = best calibrated method
                    </div>
                    """, unsafe_allow_html=True)

                # Download
                section("Download VaR Data")
                import io as _io
                var_buf = _io.StringIO()
                df_methods.to_csv(var_buf, index=False)
                st.download_button(
                    label="⬇️ Download VaR Comparison CSV",
                    data=var_buf.getvalue(),
                    file_name=f"{sym}_var_all_methods.csv",
                    mime="text/csv",
                    key=f"dl_var_{sym}",
                )


        # ── Tab 5: Residuals ──
        with main_tabs[5]:
            section("Residual Analysis — Best Model")
            if winner and winner in res and "pred" in res[winner]:
                pred_r = res[winner]["pred"]
                actual_r = sp["val"].loc[pred_r.index]
                st.plotly_chart(
                    chart_residuals(actual_r, pred_r, winner, sym),
                    use_container_width=True,
                )
                st.markdown("""
                <div class="info-box">
                    <strong>ACF of residuals:</strong> Values outside dashed bounds indicate 
                    autocorrelation — a sign the model has not captured all predictable structure.<br>
                    <strong>Q-Q plot:</strong> Deviations from the line indicate non-normality 
                    (fat tails are common in financial residuals).
                </div>
                """, unsafe_allow_html=True)
            else:
                st.info("No residuals available for the winning model.")

        # ── Tab 6: Future Forecast ──
        with main_tabs[6]:
            section(f"Future Forecast — Next {future_days} Business Days")

            if ff:
                fp  = ff["pred"]
                fl  = ff.get("lower")
                fu  = ff.get("upper")
                fm  = ff["model"]

                # Widen bands for horizon uncertainty (sqrt(h) scaling)
                best_res = test_results.get(sym, {}).get(winner, {})
                if best_res and "pred" in best_res:
                    residuals = sp["test"].loc[
                        best_res["pred"].index] - best_res["pred"]
                    res_std = float(residuals.std())
                    fl_wide, fu_wide = widen_bands_for_horizon(
                        fp, fl if fl is not None else fp,
                        fu if fu is not None else fp,
                        res_std, confidence=confidence,
                    )
                else:
                    fl_wide, fu_wide = fl, fu

                st.plotly_chart(
                    chart_future_forecast(sp["full"], fp, sym, fm,
                                         fl_wide, fu_wide),
                    use_container_width=True,
                )
                st.markdown(
                    '<div class="info-box">⚠️ <strong>Horizon uncertainty:</strong> '
                    'Prediction bands widen as √h (square-root of horizon steps). '
                    'This reflects genuine uncertainty growth — day-15 bands are '
                    f'~{np.sqrt(future_days):.1f}× wider than day-1 bands.</div>',
                    unsafe_allow_html=True,
                )

                # Forecast table
                section("Forecast Table")
                df_fc = pd.DataFrame({
                    "Date":        fp.index.strftime("%Y-%m-%d"),
                    "Forecast":    fp.values.round(2),
                    "Lower Band":  fl.values.round(2) if fl is not None else np.nan,
                    "Upper Band":  fu.values.round(2) if fu is not None else np.nan,
                    "Δ vs Last":   ((fp.values - sp["full"].iloc[-1]) / sp["full"].iloc[-1] * 100).round(2),
                }).set_index("Date")

                styled = df_fc.style.format({"Forecast":"{:.2f}","Lower Band":"{:.2f}",
                                              "Upper Band":"{:.2f}","Δ vs Last":"{:+.2f}%"})
                styled = safe_background_gradient(styled, subset=["Δ vs Last"], cmap="RdYlGn",
                                                  vmin=-5, vmax=5)
                st.dataframe(styled, use_container_width=True)

                st.download_button(
                    label="⬇️ Download Forecast CSV",
                    data=df_fc.to_csv(),
                    file_name=f"{sym}_forecast_{future_days}d.csv",
                    mime="text/csv",
                    key=f"dl_forecast_{sym}",
                )

                last_p = sp["full"].iloc[-1]
                end_p  = fp.iloc[-1]
                chg    = (end_p - last_p) / last_p * 100

                col_l, col_r, col_m = st.columns(3)
                col_l.markdown(
                    metric_card("Current Price", f"{last_p:.2f}", "", "neu"),
                    unsafe_allow_html=True,
                )
                col_r.markdown(
                    metric_card(f"Forecast +{future_days}d", f"{end_p:.2f}",
                                f"{chg:+.2f}%", "up" if chg > 0 else "down"),
                    unsafe_allow_html=True,
                )
                col_m.markdown(
                    metric_card("Model Used", fm.split("(")[0][:14],
                                badge(ff.get("type","").upper(), "blue"), "neu"),
                    unsafe_allow_html=True,
                )

                st.markdown("""
                <div class="warn-box">
                    ⚠️ <strong>Disclaimer:</strong> Forecasts are statistical model outputs and 
                    should NOT be interpreted as investment advice. Financial markets are 
                    non-stationary; model performance on historical data does not guarantee 
                    future accuracy. Always combine quantitative models with fundamental 
                    analysis and risk management.
                </div>
                """, unsafe_allow_html=True)
            else:
                st.info("Future forecast not available for this symbol.")

        # ── Tab 7: Signals ────────────────────────────────────────────────
        with main_tabs[7]:
            section("📡 Trading Signal Dashboard")

            # Compute forecast % change for signal engine
            fc_pct, fc_vol_pct = 0.0, 0.0
            if ff:
                last_p   = float(sp["full"].iloc[-1])
                fc_end_p = float(ff["pred"].iloc[-1])
                fc_pct   = (fc_end_p - last_p) / last_p * 100
                if ff.get("upper") is not None and ff.get("lower") is not None:
                    band_width = float((ff["upper"] - ff["lower"]).mean())
                    fc_vol_pct = band_width / last_p

            sig = compute_signals(
                data[sym], selected_col,
                forecast_pct=fc_pct,
                forecast_vol=fc_vol_pct,
            )

            # Signal badge
            sig_color = {"STRONG BUY":"#3FB950","BUY":"#3FB950","MILD BUY":"#56D364",
                         "HOLD":"#D29922",
                         "MILD SELL":"#F85149","SELL":"#F85149","STRONG SELL":"#FF0000"
                        }.get(sig["signal"], "#8B949E")
            st.markdown(f"""
            <div style="background:#161B22;border:2px solid {sig_color};border-radius:12px;
                        padding:24px;text-align:center;margin-bottom:20px;">
                <div style="font-family:'IBM Plex Mono',monospace;font-size:2rem;
                            font-weight:700;color:{sig_color};">{sig["signal"]}</div>
                <div style="color:#8B949E;font-size:0.85rem;margin-top:6px;">
                    Strength: {sig["strength"].upper()} &nbsp;|&nbsp;
                    Net Score: {sig["net_score"]:+d}
                    &nbsp;(Buy: {sig["buy_score"]} · Sell: {sig["sell_score"]})
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Indicator gauges
            section("Key Indicators")
            ind = sig["indicators"]
            c1,c2,c3,c4,c5 = st.columns(5)
            c1.markdown(metric_card("RSI 14", f"{ind['rsi']:.1f}",
                "Oversold" if ind['rsi']<30 else "Overbought" if ind['rsi']>70 else "Neutral",
                "up" if ind['rsi']<40 else "down" if ind['rsi']>60 else "neu"),
                unsafe_allow_html=True)
            c2.markdown(metric_card("SMA20", f"{ind['sma20']:.2f}",
                "Above SMA50 ✅" if ind['sma20']>ind['sma50'] else "Below SMA50 ⚠️",
                "up" if ind['sma20']>ind['sma50'] else "down"),
                unsafe_allow_html=True)
            c3.markdown(metric_card("SMA50", f"{ind['sma50']:.2f}",
                "Above SMA200 ✅" if ind['sma50']>ind['sma200'] else "Below SMA200 ⚠️",
                "up" if ind['sma50']>ind['sma200'] else "down"),
                unsafe_allow_html=True)
            c4.markdown(metric_card("SMA200", f"{ind['sma200']:.2f}",
                "Bull regime ✅" if ind['bull_regime'] else "Bear regime ⚠️",
                "up" if ind['bull_regime'] else "down"),
                unsafe_allow_html=True)
            c5.markdown(metric_card("Volatility 21d",
                f"{ind['vol_21d']*100:.2f}%",
                "Low ✅" if ind['vol_21d']<0.012 else "Elevated ⚠️",
                "up" if ind['vol_21d']<0.012 else "down"),
                unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # MACD & BB row
            c6,c7,c8 = st.columns(3)
            c6.markdown(metric_card("MACD",
                f"{ind['macd']:.4f}",
                f"Signal: {ind['macd_signal']:.4f}",
                "up" if ind['macd']>ind['macd_signal'] else "down"),
                unsafe_allow_html=True)
            c7.markdown(metric_card("BB Position",
                f"{ind['bb_position']:.2f}",
                "Near lower band" if ind['bb_position']<0.2 else
                "Near upper band" if ind['bb_position']>0.8 else "Mid-band",
                "up" if ind['bb_position']<0.3 else
                "down" if ind['bb_position']>0.7 else "neu"),
                unsafe_allow_html=True)
            c8.markdown(metric_card("Forecast Δ",
                f"{fc_pct:+.2f}%",
                f"Over {future_days} days",
                "up" if fc_pct>0 else "down" if fc_pct<0 else "neu"),
                unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Buy/Sell reasons
            col_b, col_s = st.columns(2)
            with col_b:
                section(f"✅ Buy Signals ({sig['buy_score']} pts)")
                for reason in sig["buy_signals"]:
                    st.markdown(
                        f'<div style="background:#1a3a22;border-left:3px solid #3FB950;'
                        f'padding:8px 12px;margin:4px 0;border-radius:0 6px 6px 0;'
                        f'font-size:0.82rem;color:#E6EDF3;">✅ {reason}</div>',
                        unsafe_allow_html=True)
                if not sig["buy_signals"]:
                    st.markdown('<div class="info-box">No active buy signals.</div>',
                                unsafe_allow_html=True)
            with col_s:
                section(f"🚨 Sell / Risk Signals ({sig['sell_score']} pts)")
                for reason in sig["sell_signals"]:
                    st.markdown(
                        f'<div style="background:#3a1a1a;border-left:3px solid #F85149;'
                        f'padding:8px 12px;margin:4px 0;border-radius:0 6px 6px 0;'
                        f'font-size:0.82rem;color:#E6EDF3;">🚨 {reason}</div>',
                        unsafe_allow_html=True)
                if not sig["sell_signals"]:
                    st.markdown('<div class="info-box">No active sell signals.</div>',
                                unsafe_allow_html=True)

            st.markdown("""
            <div class="warn-box" style="margin-top:16px;">
                ⚠️ <strong>Not investment advice.</strong> Signals are rule-based outputs
                from a quantitative model and do not account for fundamentals, news, or
                macro conditions. Always perform your own due diligence.
            </div>
            """, unsafe_allow_html=True)

        # ── Tab 8: Strategy Backtest ──────────────────────────────────────
        with main_tabs[8]:
            section("⚡ Strategy Backtesting — Forecast-Driven Long/Only")
            st.markdown("""
            <div class="info-box">
                Simulates a <strong>long-only strategy</strong>: hold when the model forecasts
                an up day, otherwise hold cash. Compares against buy-and-hold.
                <strong>No transaction costs or slippage</strong> assumed — treat as an upper bound.
            </div>
            """, unsafe_allow_html=True)

            if winner and winner in test_results.get(sym, {}):
                tr = test_results[sym][winner]
                bt = compute_strategy_backtest(
                    sp["test"], tr["pred"], confidence=confidence)

                # Summary cards
                section("Performance Summary")
                bc1,bc2,bc3,bc4 = st.columns(4)
                bc1.markdown(metric_card(
                    "Ann. Return (Strategy)",
                    f"{bt['ann_return_pct']:+.2f}%",
                    f"B&H: {bt['bh_return_pct']:+.2f}%",
                    "up" if bt['ann_return_pct'] > bt['bh_return_pct'] else "down"),
                    unsafe_allow_html=True)
                bc2.markdown(metric_card(
                    "Sharpe Ratio", f"{bt['sharpe']:.3f}",
                    "↑ >1 = good · >2 = excellent",
                    "up" if bt['sharpe']>1 else "neu" if bt['sharpe']>0 else "down"),
                    unsafe_allow_html=True)
                bc3.markdown(metric_card(
                    "Max Drawdown", f"{bt['max_drawdown_pct']:.2f}%",
                    "↑ closer to 0 is better",
                    "up" if bt['max_drawdown_pct']>-10 else "down"),
                    unsafe_allow_html=True)
                bc4.markdown(metric_card(
                    "Win Rate", f"{fmt(bt['win_rate_pct'],1,'%')}",
                    "↑ >50% = directional edge",
                    "up" if (bt['win_rate_pct'] or 0)>53 else "neu"),
                    unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                bc5,bc6,bc7,bc8 = st.columns(4)
                bc5.markdown(metric_card(
                    "Sortino Ratio", f"{bt['sortino']:.3f}",
                    "penalises downside only",
                    "up" if bt['sortino']>1 else "neu"),
                    unsafe_allow_html=True)
                bc6.markdown(metric_card(
                    "Calmar Ratio",
                    fmt(bt['calmar'],3) if not np.isnan(bt.get('calmar',np.nan)) else "N/A",
                    "Ann return / max DD",
                    "up" if (bt.get('calmar') or 0)>0.5 else "neu"),
                    unsafe_allow_html=True)
                bc7.markdown(metric_card(
                    "Profit Factor", fmt(bt['profit_factor'],3),
                    "Gross gain / gross loss",
                    "up" if (bt.get('profit_factor') or 0)>1.2 else "down"),
                    unsafe_allow_html=True)
                bc8.markdown(metric_card(
                    "# Trades", str(bt['n_trades']),
                    f"{bt['n_long_days']} long days", "neu"),
                    unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                bc9, bc10, bc11, _ = st.columns(4)
                ir_val = bt.get('information_ratio', np.nan)
                bc9.markdown(metric_card(
                    "Information Ratio",
                    fmt(ir_val, 3),
                    "Strategy excess return / tracking error",
                    "up" if (ir_val or 0) > 0.3 else "neu" if (ir_val or 0) > 0 else "down"),
                    unsafe_allow_html=True)
                hit_val = bt.get('hit_ratio_pct', np.nan)
                bc10.markdown(metric_card(
                    "Hit Ratio",
                    fmt(hit_val, 1, "%"),
                    "% active days with positive return",
                    "up" if (hit_val or 0) > 55 else "neu" if (hit_val or 0) > 50 else "down"),
                    unsafe_allow_html=True)
                bc11.markdown(metric_card(
                    "Period",
                    f"{bt['period_years']:.2f} yr",
                    "⚠️ annualised figures may be unreliable" if bt.get('annualization_warning') else "",
                    "neu"),
                    unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # Equity curve chart
                section("Equity Curve vs Buy-and-Hold")
                fig_bt = go.Figure()
                cum_bh = (1 + sp["test"].pct_change().dropna()).cumprod()
                fig_bt.add_trace(go.Scatter(
                    x=bt["cumulative"].index, y=bt["cumulative"].values,
                    name=f"Strategy ({winner.split('(')[0][:16]})",
                    line=dict(color="#3FB950", width=2)))
                fig_bt.add_trace(go.Scatter(
                    x=cum_bh.index, y=cum_bh.values,
                    name="Buy & Hold",
                    line=dict(color="#58A6FF", width=2, dash="dot")))
                fig_bt.update_layout(
                    paper_bgcolor="#0D1117", plot_bgcolor="#161B22",
                    font=dict(color="#E6EDF3", family="IBM Plex Mono"),
                    xaxis=dict(gridcolor="#30363D"),
                    yaxis=dict(gridcolor="#30363D", tickformat=".2f"),
                    legend=dict(bgcolor="#21262D", bordercolor="#30363D", borderwidth=1),
                    hovermode="x unified", margin=dict(l=60,r=20,t=40,b=40),
                    title=f"{sym} — Cumulative Return: Strategy vs Buy & Hold",
                )
                st.plotly_chart(fig_bt, use_container_width=True)

                # Drawdown chart
                section("Drawdown Series")
                fig_dd = go.Figure()
                fig_dd.add_trace(go.Scatter(
                    x=bt["drawdown_series"].index,
                    y=bt["drawdown_series"].values * 100,
                    fill="tozeroy", name="Drawdown %",
                    line=dict(color="#F85149", width=1),
                    fillcolor="rgba(248,81,73,0.2)"))
                fig_dd.update_layout(
                    paper_bgcolor="#0D1117", plot_bgcolor="#161B22",
                    font=dict(color="#E6EDF3", family="IBM Plex Mono"),
                    xaxis=dict(gridcolor="#30363D"),
                    yaxis=dict(gridcolor="#30363D", ticksuffix="%"),
                    margin=dict(l=60,r=20,t=40,b=40),
                    title=f"{sym} — Strategy Drawdown",
                )
                st.plotly_chart(fig_dd, use_container_width=True)

                # Rolling Sharpe chart
                rs_series = bt.get("rolling_sharpe_63")
                if rs_series is not None and not rs_series.dropna().empty:
                    section("Rolling 63-Day Sharpe Ratio")
                    st.plotly_chart(
                        chart_rolling_sharpe(rs_series, sym, window=63),
                        use_container_width=True,
                    )
                    st.markdown("""
                    <div class="info-box">
                        <strong>Rolling Sharpe</strong> shows how risk-adjusted performance
                        changes over time. Green fills = positive Sharpe (strategy adding value);
                        red fills = negative Sharpe (below risk-free).
                        Persistent green zones above 1.0 indicate robust, consistent performance.
                    </div>
                    """, unsafe_allow_html=True)

                # Benchmark table
                section("Benchmark Summary Table")
                bm_data = {
                    "Metric": ["Ann. Return","Sharpe","Sortino","Max Drawdown",
                               "Win Rate","Hit Ratio","Profit Factor","Calmar",
                               "Information Ratio"],
                    "Strategy": [
                        f"{bt['ann_return_pct']:+.2f}%",
                        f"{bt['sharpe']:.3f}",
                        f"{bt['sortino']:.3f}",
                        f"{bt['max_drawdown_pct']:.2f}%",
                        f"{fmt(bt['win_rate_pct'],1,'%')}",
                        f"{fmt(bt.get('hit_ratio_pct'),1,'%')}",
                        f"{fmt(bt['profit_factor'],3)}",
                        f"{fmt(bt.get('calmar'),3)}",
                        f"{fmt(bt.get('information_ratio'),3)}",
                    ],
                    "Buy & Hold": [
                        f"{bt['bh_return_pct']:+.2f}%",
                        "—","—","—","—","—","—","—","—",
                    ],
                    "Good threshold": [
                        ">10% pa",">1.0",">1.0","<-20%",">53%",">55%",">1.2",">0.5",">0.3",
                    ],
                }
                st.dataframe(pd.DataFrame(bm_data), use_container_width=True, hide_index=True)
            else:
                st.info("Run the analysis first — test results needed for backtesting.")
# ─────────────────────────────────────────────────────────────────────────────

def _fig_div(fig, height: int = 420) -> str:
    """Convert a Plotly figure to an HTML div string (no full page wrapper)."""
    import plotly.io as pio
    fig.update_layout(height=height, margin=dict(l=50, r=20, t=40, b=40))
    return pio.to_html(fig, full_html=False, include_plotlyjs=False,
                       config={"displayModeBar": True, "scrollZoom": False})


def _metric_html(label: str, value: str, delta: str = "",
                 color: str = "#E6EDF3") -> str:
    return f"""
    <div class="mc">
      <div class="mc-label">{label}</div>
      <div class="mc-value" style="color:{color}">{value}</div>
      {"<div class='mc-delta'>" + delta + "</div>" if delta else ""}
    </div>"""


def _table_html(headers: list, rows: list, col_colors: dict = None) -> str:
    """Render a styled HTML table. col_colors: {col_idx: fn(val)->color}."""
    th = "".join(f"<th>{h}</th>" for h in headers)
    body = ""
    for ri, row in enumerate(rows):
        cls = "even" if ri % 2 == 0 else "odd"
        tds = ""
        for ci, v in enumerate(row):
            color = ""
            if col_colors and ci in col_colors:
                color = f'style="color:{col_colors[ci](v)};font-weight:600"'
            tds += f"<td {color}>{v}</td>"
        body += f"<tr class='{cls}'>{tds}</tr>"
    return f"<table><thead><tr>{th}</tr></thead><tbody>{body}</tbody></table>"


def build_html_report(
    loaded_syms, splits, all_results, test_results,
    final_winners, wf_results, var_results, future_forecasts,
    selected_col, future_days, var_best_method, var_conf,
    top_n_compare,
) -> bytes:
    from core.charts import (
        chart_price, chart_forecast, chart_model_comparison,
        chart_var_breaches, chart_walk_forward, chart_directional_accuracy,
        chart_future_forecast, chart_returns_dist, chart_residuals,
    )
    import plotly.io as pio

    # ── CSS ──────────────────────────────────────────────────────────────────
    CSS = """
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=Space+Grotesk:wght@400;500;600;700&display=swap');
    *{box-sizing:border-box;margin:0;padding:0}
    body{background:#0D1117;color:#E6EDF3;font-family:'Space Grotesk',sans-serif;
         font-size:14px;padding:0 0 60px 0}
    a{color:#58A6FF}

    /* ── Header ── */
    .report-header{background:linear-gradient(135deg,#161B22,#0D1117);
      border-bottom:2px solid #21262D;padding:36px 48px 28px;
      display:flex;justify-content:space-between;align-items:flex-end}
    .report-title{font-family:'IBM Plex Mono',monospace;font-size:2rem;
      font-weight:700;color:#58A6FF;letter-spacing:0.04em}
    .report-sub{font-size:0.78rem;color:#8B949E;letter-spacing:0.12em;
      text-transform:uppercase;margin-top:6px;font-family:'IBM Plex Mono',monospace}
    .report-meta{text-align:right;font-family:'IBM Plex Mono',monospace;
      font-size:0.72rem;color:#8B949E;line-height:1.8}

    /* ── Symbol section ── */
    .sym-header{background:#161B22;border-top:3px solid #58A6FF;
      border-bottom:1px solid #30363D;padding:20px 48px 16px;
      margin-top:48px;display:flex;align-items:baseline;gap:16px}
    .sym-title{font-family:'IBM Plex Mono',monospace;font-size:1.5rem;
      font-weight:700;color:#E6EDF3}
    .sym-badge{font-family:'IBM Plex Mono',monospace;font-size:0.70rem;
      letter-spacing:0.12em;text-transform:uppercase;color:#8B949E}

    /* ── Content wrapper ── */
    .content{padding:0 48px}

    /* ── Section title ── */
    .section{font-family:'IBM Plex Mono',monospace;font-size:0.68rem;
      letter-spacing:0.20em;text-transform:uppercase;color:#58A6FF;
      border-bottom:1px solid #30363D;padding-bottom:6px;
      margin:32px 0 16px 0}

    /* ── Metric cards ── */
    .metrics-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));
      gap:12px;margin-bottom:24px}
    .mc{background:#161B22;border:1px solid #30363D;border-radius:8px;
      padding:14px 16px}
    .mc-label{font-family:'IBM Plex Mono',monospace;font-size:0.64rem;
      letter-spacing:0.12em;text-transform:uppercase;color:#8B949E;margin-bottom:6px}
    .mc-value{font-family:'IBM Plex Mono',monospace;font-size:1.35rem;
      font-weight:600;color:#E6EDF3;line-height:1.1}
    .mc-delta{font-size:0.72rem;color:#8B949E;margin-top:4px;
      font-family:'IBM Plex Mono',monospace}

    /* ── Chart container ── */
    .chart-wrap{background:#161B22;border:1px solid #30363D;border-radius:8px;
      padding:8px;margin-bottom:20px;overflow:hidden}
    .chart-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
    .chart-row .chart-wrap{margin-bottom:0}

    /* ── Tables ── */
    table{width:100%;border-collapse:collapse;margin-bottom:20px;font-size:0.82rem}
    thead tr{background:#21262D}
    th{font-family:'IBM Plex Mono',monospace;font-size:0.68rem;letter-spacing:0.08em;
       text-transform:uppercase;color:#8B949E;padding:10px 12px;
       text-align:left;border-bottom:1px solid #30363D}
    td{padding:9px 12px;border-bottom:1px solid #21262D;color:#E6EDF3}
    tr.even{background:#0D1117}
    tr.odd{background:#161B22}
    tr:hover td{background:#21262D}

    /* ── Info / warn boxes ── */
    .info-box{background:#1a2a3a;border-left:3px solid #58A6FF;
      padding:10px 16px;border-radius:0 6px 6px 0;
      font-size:0.84rem;margin:12px 0;line-height:1.6}
    .warn-box{background:#3a2a1a;border-left:3px solid #D29922;
      padding:10px 16px;border-radius:0 6px 6px 0;
      font-size:0.84rem;margin:12px 0}

    /* ── TOC ── */
    .toc{background:#161B22;border:1px solid #30363D;border-radius:8px;
      padding:20px 28px;margin:32px 48px;display:flex;gap:32px;flex-wrap:wrap}
    .toc-title{font-family:'IBM Plex Mono',monospace;font-size:0.68rem;
      letter-spacing:0.15em;text-transform:uppercase;color:#58A6FF;
      margin-bottom:10px;width:100%}
    .toc a{color:#8B949E;text-decoration:none;font-size:0.82rem;
      font-family:'IBM Plex Mono',monospace;display:block;margin-bottom:4px}
    .toc a:hover{color:#58A6FF}

    /* ── Footer ── */
    .report-footer{border-top:1px solid #30363D;margin-top:64px;
      padding:20px 48px;text-align:center;
      font-family:'IBM Plex Mono',monospace;font-size:0.65rem;
      color:#8B949E;letter-spacing:0.10em}

    /* ── Page break for print ── */
    @media print{.sym-header{page-break-before:always}}
    """

    # ── Plotly.js (CDN, loaded once) ─────────────────────────────────────────
    PLOTLY_JS = '<script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>'

    now_str  = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    syms_str = ", ".join(loaded_syms)

    # ── TOC ──────────────────────────────────────────────────────────────────
    toc_links = ""
    for sym in loaded_syms:
        toc_links += f'<a href="#{sym}">📈 {sym}</a>'

    toc_html = f"""
    <div class="toc">
      <div class="toc-title">Contents</div>
      {toc_links}
    </div>"""

    # ── Per-symbol sections ───────────────────────────────────────────────────
    body_parts = []

    for sym in loaded_syms:
        sp     = splits[sym]
        res    = all_results[sym]
        t_res  = test_results.get(sym, {})
        winner = final_winners.get(sym)
        w_res  = wf_results.get(sym, [])
        all_vr_sym  = var_results.get(sym, {})
        best_vm_sym = var_best_method.get(sym, "historical") if isinstance(var_best_method, dict) else "historical"
        vr     = all_vr_sym.get(best_vm_sym, {})
        ff     = future_forecasts.get(sym, {})
        bt     = t_res.get(winner, {}) if winner else {}

        def _f(v, d=2, s=""):
            return f"{v:.{d}f}{s}" if v is not None and not (isinstance(v, float) and np.isnan(v)) else "N/A"

        # ── Metric cards ─────────────────────────────────────────────────────
        da_val  = bt.get("da", 0) or 0
        da_col  = "#3FB950" if da_val > 55 else ("#F85149" if da_val < 50 else "#E6EDF3")
        fc_end  = float(ff["pred"].iloc[-1]) if ff else None
        last_p  = float(sp["full"].iloc[-1])
        delta_p = ((fc_end - last_p) / last_p * 100) if fc_end else None

        cards = "".join([
            _metric_html("Best Model",
                         winner.split("(")[0][:16] if winner else "—",
                         res.get(winner,{}).get("type","").upper() if winner else ""),
            _metric_html("Test MAPE", _f(bt.get("mape"), 2, "%"),
                         "↓ lower = better"),
            _metric_html("Test RMSE", _f(bt.get("rmse"), 2)),
            _metric_html("Dir. Accuracy", _f(da_val, 1, "%"),
                         "↑ above 50% = edge", da_col),
            _metric_html("VaR Breaches",
                         f"{vr.get('n_breaches','?')}/{vr.get('n_obs','?')}",
                         f"exp: {_f(vr.get('exp_rate',0)*100,1)}%"),
            _metric_html(f"Forecast +{future_days}d",
                         _f(fc_end, 2) if fc_end else "N/A",
                         (f"{delta_p:+.2f}%" if delta_p is not None else ""),
                         "#3FB950" if (delta_p or 0) > 0 else "#F85149"),
            _metric_html("Last Price", _f(last_p, 2)),
            _metric_html("Train Bars", str(len(sp["train"]))),
            _metric_html("Val Bars",   str(len(sp["val"]))),
            _metric_html("Test Bars",  str(len(sp["test"]))),
        ])

        # ── Charts ───────────────────────────────────────────────────────────
        # 1. Price split
        try:
            c_price = _fig_div(chart_price(
                sp["train"], sp["val"], sp["test"], sym, selected_col), 380)
        except Exception:
            c_price = "<p style='color:#8B949E'>Price chart unavailable</p>"

        # 2. Validation forecast
        try:
            c_val_fc = _fig_div(chart_forecast(
                sp["val"], res, sym, top_n=top_n_compare,
                show_bands=True, history=sp["train"]), 400)
        except Exception:
            c_val_fc = "<p style='color:#8B949E'>Val forecast unavailable</p>"

        # 3. Test forecast
        try:
            c_test_fc = _fig_div(chart_forecast(
                sp["test"], t_res, sym, top_n=top_n_compare,
                show_bands=True,
                history=pd.concat([sp["train"], sp["val"]])), 400) if t_res else ""
        except Exception:
            c_test_fc = ""

        # 4. Model comparison
        try:
            c_model_cmp = _fig_div(chart_model_comparison(res, sym), 380)
        except Exception:
            c_model_cmp = ""

        # 5. Directional accuracy
        try:
            c_da = _fig_div(chart_directional_accuracy({**res, **t_res}, sym), 360)
        except Exception:
            c_da = ""

        # 6. Walk-forward CV
        try:
            c_wf = _fig_div(chart_walk_forward(w_res, sym), 420) if w_res else ""
        except Exception:
            c_wf = ""

        # 7. VaR breach chart
        try:
            c_var = _fig_div(chart_var_breaches(
                vr["returns"], vr["var_series"],
                vr["breaches"], sym, var_conf), 380) if vr and "returns" in vr else ""
        except Exception:
            c_var = ""

        # 8. Returns distribution
        try:
            if vr and "returns" in vr:
                from core.risk import var_historical
                ret     = vr["returns"]
                var95   = float(vr["var_series"].mean())
                var99s  = var_historical(ret, 0.99, window=252).dropna()
                var99   = float(var99s.mean()) if len(var99s) else 0
                c_rdist = _fig_div(chart_returns_dist(ret, sym, var95, var99), 380)
            else:
                c_rdist = ""
        except Exception:
            c_rdist = ""

        # 9. Residuals
        try:
            if winner and winner in res and "pred" in res[winner]:
                pred_r  = res[winner]["pred"]
                act_r   = sp["val"].loc[pred_r.index]
                c_resid = _fig_div(chart_residuals(act_r, pred_r, winner, sym), 480)
            else:
                c_resid = ""
        except Exception:
            c_resid = ""

        # 10. Future forecast
        try:
            if ff:
                c_future = _fig_div(chart_future_forecast(
                    sp["full"], ff["pred"], sym,
                    ff["model"], ff.get("lower"), ff.get("upper")), 380)
            else:
                c_future = ""
        except Exception:
            c_future = ""

        # ── Validation metrics table ──────────────────────────────────────────
        val_rows = []
        for name, r in sorted(
            [(n, r) for n, r in res.items()
             if not n.startswith("__") and isinstance(r, dict) and r.get("mape") is not None],
            key=lambda x: x[1]["mape"]
        ):
            val_rows.append([
                name, r.get("type",""),
                _f(r.get("mape"), 3, "%"),
                _f(r.get("rmse"), 3),
                _f(r.get("mae"),  3),
                _f(r.get("da"),   1, "%"),
                _f(r.get("train_mape"), 3, "%"),
                _f(r.get("overfit_score"), 3),
            ])

        def _da_color(v):
            try:
                n = float(str(v).replace("%",""))
                return "#3FB950" if n > 55 else ("#F85149" if n < 50 else "#E6EDF3")
            except Exception:
                return "#E6EDF3"

        val_table = _table_html(
            ["Model","Type","Val MAPE %","RMSE","MAE","DA %","Train MAPE %","Overfit"],
            val_rows, col_colors={5: _da_color})

        # ── Test metrics table ────────────────────────────────────────────────
        test_rows = []
        for name, r in sorted(t_res.items(), key=lambda x: x[1].get("mape", 999)):
            test_rows.append([
                ("⭐ " if name == winner else "") + name,
                r.get("type",""),
                _f(r.get("mape"), 3, "%"),
                _f(r.get("rmse"), 3),
                _f(r.get("mae"),  3),
                _f(r.get("da"),   1, "%"),
            ])
        test_table = _table_html(
            ["Model","Type","Test MAPE %","RMSE","MAE","DA %"],
            test_rows, col_colors={5: _da_color}) if test_rows else ""

        # ── Walk-forward table ────────────────────────────────────────────────
        wf_rows = []
        for fold in w_res:
            wf_rows.append([
                fold.get("fold"),
                fold.get("train_size"),
                fold.get("test_size"),
                _f(fold.get("mape"), 3, "%"),
                _f(fold.get("rmse"), 3),
                _f(fold.get("mae"),  3),
                _f(fold.get("da"),   1, "%"),
            ])
        if wf_rows:
            mapes_wf = [fold.get("mape", np.nan) for fold in w_res]
            das_wf   = [fold.get("da",   np.nan) for fold in w_res]
            wf_rows.append([
                "MEAN ± STD", "", "",
                f"{np.nanmean(mapes_wf):.3f} ± {np.nanstd(mapes_wf):.3f}",
                "", "",
                f"{np.nanmean(das_wf):.1f} ± {np.nanstd(das_wf):.1f}",
            ])
        wf_table = _table_html(
            ["Fold","Train","Test","MAPE %","RMSE","MAE","DA %"],
            wf_rows) if wf_rows else ""

        # ── VaR stats table ───────────────────────────────────────────────────
        if vr and "kupiec" in vr:
            kup   = vr["kupiec"]
            chri  = vr["christoffersen"]
            binom = vr["binomial_uc"]

            def _pv_color(v):
                try:
                    p = float(str(v))
                    return "#3FB950" if p >= 0.05 else "#F85149"
                except Exception:
                    return "#E6EDF3"

            var_summary_rows = [
                ["Kupiec POF",
                 _f(kup.get("lr_stat"),3),
                 _f(kup.get("p_value"),4),
                 "✓ Pass" if not kup.get("reject_h0") else "✗ Fail",
                 _f(kup.get("obs_rate",0)*100,2,"%"),
                 _f(kup.get("exp_rate",0)*100,2,"%"),
                 f"{kup.get('n_breaches','?')}/{kup.get('n','?')}"],
                ["Christoffersen CC",
                 _f(chri.get("lr_cc"),3),
                 _f(chri.get("p_value_cc"),4),
                 "✓ Pass" if not chri.get("reject_h0_cc") else "✗ Fail",
                 "—","—","—"],
                ["Binomial UC",
                 "—",
                 _f(binom.get("p_value"),4),
                 "✓ Pass" if not binom.get("reject_h0") else "✗ Fail",
                 _f(binom.get("obs_rate",0)*100,2,"%"),
                 _f(binom.get("exp_rate",0)*100,2,"%"),
                 f"{binom.get('n_breaches','?')}/{binom.get('n','?')}"],
            ]

            def _pass_color(v):
                return "#3FB950" if "Pass" in str(v) else "#F85149"

            var_table = _table_html(
                ["Test","LR Stat","p-value","Result",
                 "Obs Rate","Exp Rate","Breaches"],
                var_summary_rows,
                col_colors={2: _pv_color, 3: _pass_color})
            es_val = _f(float(vr.get("es", np.nan))*100, 3, "%")
        else:
            var_table = ""
            es_val = "N/A"

        # ── Forecast table ────────────────────────────────────────────────────
        if ff:
            fp = ff["pred"]; fl = ff.get("lower"); fu = ff.get("upper")
            fc_rows = []
            for dt, fv in fp.items():
                lo = _f(float(fl.loc[dt]), 2) if fl is not None and dt in fl.index else "—"
                hi = _f(float(fu.loc[dt]), 2) if fu is not None and dt in fu.index else "—"
                d  = float(fv) - last_p
                dp = d / last_p * 100
                fc_rows.append([
                    str(dt.date()), _f(float(fv),2),
                    lo, hi, f"{d:+.2f}", f"{dp:+.2f}%"])
            def _delta_color(v):
                try:
                    return "#3FB950" if float(str(v).replace("%","")) >= 0 else "#F85149"
                except Exception:
                    return "#E6EDF3"
            fc_table = _table_html(
                ["Date","Forecast","Lower","Upper","Δ Price","Δ %"],
                fc_rows, col_colors={4: _delta_color, 5: _delta_color})
        else:
            fc_table = ""

        # ── Assemble symbol section ───────────────────────────────────────────
        section_html = f"""
<div id="{sym}">
  <div class="sym-header">
    <span class="sym-title">📈 {sym}</span>
    <span class="sym-badge">{selected_col} · {sp['full'].index[0].date()} → {sp['full'].index[-1].date()}</span>
  </div>
  <div class="content">

    <div class="section">Summary Metrics</div>
    <div class="metrics-row">{cards}</div>

    <div class="section">Price History — Train / Val / Test</div>
    <div class="chart-wrap">{c_price}</div>

    <div class="section">Validation Forecast — Top {top_n_compare} Models</div>
    <div class="chart-wrap">{c_val_fc}</div>

    {"<div class='section'>Test Forecast — Top Models</div><div class='chart-wrap'>" + c_test_fc + "</div>" if c_test_fc else ""}

    <div class="section">Model Performance — Validation Set</div>
    {val_table}

    {"<div class='section'>Model Performance — Test Set</div>" + test_table if test_table else ""}

    <div class="section">Model Comparison</div>
    <div class="chart-row">
      <div class="chart-wrap">{c_model_cmp}</div>
      <div class="chart-wrap">{c_da}</div>
    </div>

    {"<div class='section'>Walk-Forward Cross-Validation (5 Folds)</div><div class='chart-wrap'>" + c_wf + "</div>" + wf_table if c_wf else ""}

    {"<div class='section'>VaR Backtest — " + best_vm_sym.upper() + " " + str(int(var_conf*100)) + "% CI</div><div class='chart-wrap'>" + c_var + "</div>" if c_var else ""}
    {"<div class='info-box'>Expected Shortfall (CVaR): <strong>" + es_val + "</strong> average loss beyond VaR threshold</div>" if vr and "kupiec" in vr else ""}
    {var_table}

    {"<div class='section'>Returns Distribution & Normality</div><div class='chart-wrap'>" + c_rdist + "</div>" if c_rdist else ""}

    {"<div class='section'>Residual Analysis — " + (winner or "") + "</div><div class='chart-wrap'>" + c_resid + "</div>" if c_resid else ""}

    {"<div class='section'>Future Forecast — Next " + str(future_days) + " Business Days</div><div class='chart-wrap'>" + c_future + "</div>" + fc_table if c_future else ""}

    <div class="warn-box">
      ⚠️ <strong>Disclaimer:</strong> Forecasts are statistical model outputs for research
      purposes only and should NOT be interpreted as investment advice.
    </div>
  </div>
</div>"""
        body_parts.append(section_html)

    # ── Assemble full HTML document ───────────────────────────────────────────
    all_syms_str = " · ".join(loaded_syms)
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>QuantLens Report — {syms_str}</title>
{PLOTLY_JS}
<style>{CSS}</style>
</head>
<body>

<div class="report-header">
  <div>
    <div class="report-title">📈 QuantLens</div>
    <div class="report-sub">Multi-Model Stock Forecasting & Risk Analytics Report</div>
  </div>
  <div class="report-meta">
    Symbols: {syms_str}<br>
    Price Column: {selected_col}<br>
    Generated: {now_str}<br>
    Forecast Horizon: {future_days}d · VaR: Auto-selected {int(var_conf*100)}%
  </div>
</div>

{toc_html}

{"".join(body_parts)}

<div class="report-footer">
  QuantLens · Multi-Model Stock Forecasting System ·
  ARIMA / Log-ARIMA / Holt-Winters / XGBoost / GBM / RandomForest / LSTM / GRU ·
  Walk-Forward CV · Kupiec + Christoffersen VaR Tests · For research &amp; educational purposes only.
</div>

</body>
</html>"""

    return html.encode("utf-8")


# ── Report download button ────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("""
<div style="background:linear-gradient(135deg,#161B22,#1C2128);
            border:1px solid #30363D;border-radius:12px;
            padding:24px 32px;margin:8px 0 28px 0;">
  <div style="display:flex;align-items:center;gap:18px;margin-bottom:16px;">
    <div style="font-size:2rem;">📥</div>
    <div>
      <div style="font-family:'IBM Plex Mono',monospace;font-size:1.05rem;
                  font-weight:700;color:#E6EDF3;">Download Complete Analysis Report</div>
      <div style="font-size:0.82rem;color:#8B949E;margin-top:4px;">
        Single HTML file · All interactive charts · Full metrics tables ·
        VaR tests · Forecasts · Opens in any browser
      </div>
    </div>
  </div>
  <div style="display:flex;gap:10px;flex-wrap:wrap;">
    <span style="background:#1a2a3a;border:1px solid #58A6FF;border-radius:20px;
                 padding:3px 12px;font-size:0.70rem;font-family:IBM Plex Mono,monospace;
                 color:#58A6FF;">📊 10+ Interactive Charts</span>
    <span style="background:#1a3a22;border:1px solid #3FB950;border-radius:20px;
                 padding:3px 12px;font-size:0.70rem;font-family:IBM Plex Mono,monospace;
                 color:#3FB950;">📋 All Model Metrics</span>
    <span style="background:#3a2a1a;border:1px solid #D29922;border-radius:20px;
                 padding:3px 12px;font-size:0.70rem;font-family:IBM Plex Mono,monospace;
                 color:#D29922;">🛡️ VaR Backtest Results</span>
    <span style="background:#2a1a3a;border:1px solid #BC8CFF;border-radius:20px;
                 padding:3px 12px;font-size:0.70rem;font-family:IBM Plex Mono,monospace;
                 color:#BC8CFF;">🔮 Future Forecasts</span>
  </div>
</div>
""", unsafe_allow_html=True)

with st.spinner("Building full report — rendering all charts…"):
    try:
        report_bytes = build_html_report(
            loaded_syms, splits, all_results, test_results,
            final_winners, wf_results, var_results, future_forecasts,
            selected_col, future_days, var_best_method, var_conf,
            top_n_compare,
        )
        fname = (f"QuantLens_Report_{'_'.join(loaded_syms)}_"
                 f"{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.html")
        st.download_button(
            label="⬇️  Download Full Report (.html)",
            data=report_bytes,
            file_name=fname,
            mime="text/html",
            use_container_width=True,
        )
        size_kb = len(report_bytes) // 1024
        st.caption(f"Report ready · {size_kb:,} KB · Open in Chrome / Firefox / Edge")
    except Exception as _rpt_e:
        st.warning(f"Report generation failed: {_rpt_e}")

def build_report_excel(
    loaded_syms, splits, all_results, test_results,
    final_winners, wf_results, var_results, future_forecasts,
    selected_col, future_days, var_best_method, var_conf,
) -> bytes:
    """Build a fully-formatted Excel report and return as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    # ── Palette ──────────────────────────────────────────────────────────────
    C_BG      = "0D1117"   # dark bg
    C_HEADER  = "161B22"   # surface
    C_ACCENT  = "58A6FF"   # blue
    C_GREEN   = "3FB950"
    C_RED     = "F85149"
    C_ORANGE  = "D29922"
    C_PURPLE  = "BC8CFF"
    C_TEXT    = "E6EDF3"
    C_MUTED   = "8B949E"
    C_WHITE   = "FFFFFF"

    def hdr_font(color=C_TEXT, bold=True, sz=11):
        return Font(name="Arial", bold=bold, color=color, size=sz)

    def hdr_fill(color=C_HEADER):
        return PatternFill("solid", fgColor=color)

    def thin_border():
        s = Side(style="thin", color="30363D")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def right():
        return Alignment(horizontal="right", vertical="center")

    def apply_header_row(ws, row_idx, labels, col_colors=None):
        """Write a styled header row."""
        for ci, label in enumerate(labels, 1):
            c = ws.cell(row=row_idx, column=ci, value=label)
            c.font      = hdr_font(color=C_TEXT, bold=True, sz=10)
            c.fill      = hdr_fill(C_HEADER)
            c.alignment = center()
            c.border    = thin_border()

    def write_data_row(ws, row_idx, values, num_fmt=None):
        """Write a data row with alternating shading."""
        fill_color = "161B22" if row_idx % 2 == 0 else "1C2128"
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font      = Font(name="Arial", color=C_TEXT, size=10)
            c.fill      = PatternFill("solid", fgColor=fill_color)
            c.alignment = right() if isinstance(val, (int, float)) else center()
            c.border    = thin_border()
            if num_fmt and ci <= len(num_fmt) and num_fmt[ci-1]:
                c.number_format = num_fmt[ci-1]

    def title_block(ws, title, subtitle=""):
        """Write a branded title block at the top of a sheet."""
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = C_ACCENT
        # Row 1 — big title
        ws.merge_cells("A1:L1")
        c = ws["A1"]
        c.value     = f"QuantLens · {title}"
        c.font      = Font(name="Arial", bold=True, color=C_ACCENT, size=14)
        c.fill      = hdr_fill(C_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        # Row 2 — subtitle
        ws.merge_cells("A2:L2")
        c = ws["A2"]
        c.value     = subtitle or f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
        c.font      = Font(name="Arial", color=C_MUTED, size=9)
        c.fill      = hdr_fill(C_BG)
        ws.row_dimensions[2].height = 16

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def color_cell(ws, row, col, bg=None, fg=C_TEXT, bold=False):
        c = ws.cell(row=row, column=col)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", color=fg, bold=bold, size=10)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — CROSS-SYMBOL SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = C_ACCENT
    title_block(ws, "Analysis Summary",
                f"Symbols: {', '.join(loaded_syms)} | Price column: {selected_col}")

    hdrs = ["Symbol", "Best Model", "Type",
            "Val MAPE %", "Test MAPE %", "Test RMSE", "Test MAE",
            "Dir. Acc %", "VaR Method", "VaR Breaches",
            "Obs Rate %", "Exp Rate %", "ES (CVaR %)",
            f"Forecast +{future_days}d", "Δ vs Last %"]
    apply_header_row(ws, 4, hdrs)

    for ri, sym in enumerate(loaded_syms, 5):
        winner   = final_winners.get(sym)
        t_res    = test_results.get(sym, {})
        bt       = t_res.get(winner, {}) if winner else {}
        _best_vm = var_best_method.get(sym, "historical") if isinstance(var_best_method, dict) else "historical"
        vr       = var_results.get(sym, {}).get(_best_vm, {})
        ff       = future_forecasts.get(sym, {})
        val_mape = all_results[sym].get(winner, {}).get("mape", np.nan) if winner else np.nan
        last_p   = splits[sym]["full"].iloc[-1]
        fc_end   = float(ff["pred"].iloc[-1]) if ff else np.nan
        delta_pct = (fc_end - last_p) / last_p * 100 if ff else np.nan
        row = [
            sym,
            winner.split("(")[0][:24] if winner else "—",
            all_results[sym].get(winner, {}).get("type", "—") if winner else "—",
            round(val_mape, 3) if not np.isnan(val_mape) else "—",
            round(bt.get("mape", np.nan), 3) if bt else "—",
            round(bt.get("rmse", np.nan), 3) if bt else "—",
            round(bt.get("mae",  np.nan), 3) if bt else "—",
            round(bt.get("da",   np.nan), 1) if bt else "—",
            _best_vm,
            f"{vr.get('n_breaches','?')} / {vr.get('n_obs','?')}",
            round(vr.get("obs_rate", np.nan)*100, 2) if vr else "—",
            round(vr.get("exp_rate", np.nan)*100, 2) if vr else "—",
            round(float(vr.get("es", np.nan))*100, 3) if vr and not np.isnan(vr.get("es", np.nan)) else "—",
            round(fc_end, 2) if not np.isnan(fc_end) else "—",
            round(delta_pct, 2) if not np.isnan(delta_pct) else "—",
        ]
        write_data_row(ws, ri, row)
        # Colour DA cell
        da_val = bt.get("da", 0) if bt else 0
        da_col = 8
        if da_val > 55:
            color_cell(ws, ri, da_col, fg=C_GREEN, bold=True)
        elif da_val < 50:
            color_cell(ws, ri, da_col, fg=C_RED)
        # Colour delta cell
        if ff and not np.isnan(delta_pct):
            color_cell(ws, ri, 15, fg=(C_GREEN if delta_pct > 0 else C_RED), bold=True)

    set_col_widths(ws, [10,26,14,11,11,11,11,11,11,14,11,11,11,14,11])
    ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEETS 2..N — PER-SYMBOL MODEL METRICS
    # ══════════════════════════════════════════════════════════════════════════
    for sym in loaded_syms:
        ws2 = wb.create_sheet(f"{sym} Models")
        ws2.sheet_properties.tabColor = C_PURPLE
        title_block(ws2, f"{sym} — Model Performance",
                    f"Validation & Test metrics for all trained models")

        # Val section
        ws2["A4"] = "VALIDATION SET"
        ws2["A4"].font = hdr_font(color=C_ACCENT, sz=10)
        ws2["A4"].fill = hdr_fill(C_BG)

        v_hdrs = ["Model", "Type", "Val MAPE %", "Val RMSE", "Val MAE",
                  "DA %", "Train MAPE %", "Overfit Score"]
        apply_header_row(ws2, 5, v_hdrs)

        val_rows = []
        for name, r in all_results[sym].items():
            if name.startswith("__") or not isinstance(r, dict) or r.get("mape") is None:
                continue
            val_rows.append([
                name, r.get("type",""),
                round(r.get("mape", np.nan), 3),
                round(r.get("rmse", np.nan), 3),
                round(r.get("mae",  np.nan), 3),
                round(r.get("da",   np.nan), 1),
                round(r.get("train_mape", np.nan), 3),
                round(r.get("overfit_score", np.nan), 3),
            ])
        val_rows.sort(key=lambda x: x[2] if isinstance(x[2], float) else 999)
        for ri, row in enumerate(val_rows, 6):
            write_data_row(ws2, ri, row)

        # Test section
        t_start = 6 + len(val_rows) + 2
        ws2.cell(row=t_start, column=1, value="TEST SET").font = hdr_font(color=C_GREEN, sz=10)
        ws2.cell(row=t_start, column=1).fill = hdr_fill(C_BG)

        t_hdrs = ["Model", "Type", "Test MAPE %", "Test RMSE",
                  "Test MAE", "DA %"]
        apply_header_row(ws2, t_start+1, t_hdrs)

        t_res = test_results.get(sym, {})
        t_rows = []
        for name, r in t_res.items():
            t_rows.append([
                name, r.get("type",""),
                round(r.get("mape", np.nan), 3),
                round(r.get("rmse", np.nan), 3),
                round(r.get("mae",  np.nan), 3),
                round(r.get("da",   np.nan), 1),
            ])
        t_rows.sort(key=lambda x: x[2] if isinstance(x[2], float) else 999)
        for ri, row in enumerate(t_rows, t_start+2):
            write_data_row(ws2, ri, row)
            winner = final_winners.get(sym)
            if row[0] == winner:
                for ci in range(1, 7):
                    ws2.cell(row=ri, column=ci).font = Font(
                        name="Arial", color=C_GREEN, bold=True, size=10)

        set_col_widths(ws2, [30, 14, 12, 12, 12, 10, 12, 12])
        ws2.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET — WALK-FORWARD CV
    # ══════════════════════════════════════════════════════════════════════════
    wf_ws = wb.create_sheet("Walk-Forward CV")
    wf_ws.sheet_properties.tabColor = C_ORANGE
    title_block(wf_ws, "Walk-Forward Cross-Validation",
                "Expanding-window CV (5 folds) · Reference model: AutoARIMA")

    row_cursor = 4
    for sym in loaded_syms:
        w_res = wf_results.get(sym, [])
        if not w_res:
            continue
        wf_ws.cell(row=row_cursor, column=1, value=sym).font = hdr_font(color=C_ACCENT, sz=11)
        wf_ws.cell(row=row_cursor, column=1).fill = hdr_fill(C_BG)
        row_cursor += 1

        hdrs_wf = ["Fold", "Train Size", "Test Size",
                   "MAPE %", "RMSE", "MAE", "DA %"]
        apply_header_row(wf_ws, row_cursor, hdrs_wf)
        row_cursor += 1

        mapes, das = [], []
        for fold in w_res:
            row = [
                fold.get("fold"),
                fold.get("train_size"),
                fold.get("test_size"),
                round(fold.get("mape", np.nan), 3),
                round(fold.get("rmse", np.nan), 3),
                round(fold.get("mae",  np.nan), 3),
                round(fold.get("da",   np.nan), 1),
            ]
            write_data_row(wf_ws, row_cursor, row)
            mapes.append(fold.get("mape", np.nan))
            das.append(fold.get("da", np.nan))
            row_cursor += 1

        # Summary row
        summary = ["MEAN ± STD", "", "",
                   f"{np.nanmean(mapes):.3f} ± {np.nanstd(mapes):.3f}",
                   "", "",
                   f"{np.nanmean(das):.1f} ± {np.nanstd(das):.1f}"]
        for ci, val in enumerate(summary, 1):
            c = wf_ws.cell(row=row_cursor, column=ci, value=val)
            c.font = Font(name="Arial", color=C_ORANGE, bold=True, size=10)
            c.fill = hdr_fill(C_BG)
            c.border = thin_border()
        row_cursor += 2

    set_col_widths(wf_ws, [8, 12, 12, 12, 12, 12, 10])
    wf_ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET — VAR BACKTEST
    # ══════════════════════════════════════════════════════════════════════════
    var_ws = wb.create_sheet("VaR Backtest")
    var_ws.sheet_properties.tabColor = C_RED
    title_block(var_ws, "VaR Backtesting Results",
                f"Auto-selected best method per symbol | Confidence: {var_conf*100:.0f}%")

    # Summary table
    apply_header_row(var_ws, 4, [
        "Symbol", "Method", "Confidence",
        "N Obs", "N Breaches", "Obs Rate %", "Exp Rate %",
        "ES (CVaR %)",
        "Kupiec p", "Kupiec Pass",
        "Christoffersen p", "CC Pass",
        "Binomial p", "Binomial Pass",
    ])

    for ri, sym in enumerate(loaded_syms, 5):
        vr = var_results.get(sym, {})
        if not vr or "error" in vr:
            write_data_row(var_ws, ri, [sym, "—"*13])
            continue
        kup  = vr.get("kupiec", {})
        chri = vr.get("christoffersen", {})
        binom = vr.get("binomial_uc", {})
        row = [
            sym,
            vr.get("var_method", _best_vm),
            f"{var_conf*100:.0f}%",
            vr.get("n_obs", "—"),
            vr.get("n_breaches", "—"),
            round(vr.get("obs_rate", np.nan)*100, 3),
            round(vr.get("exp_rate", np.nan)*100, 3),
            round(float(vr.get("es", np.nan))*100, 4) if not np.isnan(vr.get("es", np.nan)) else "—",
            round(kup.get("p_value", np.nan), 4),
            "✓ Pass" if not kup.get("reject_h0", True) else "✗ Fail",
            round(chri.get("p_value_cc", np.nan), 4),
            "✓ Pass" if not chri.get("reject_h0_cc", True) else "✗ Fail",
            round(binom.get("p_value", np.nan), 4),
            "✓ Pass" if not binom.get("reject_h0", True) else "✗ Fail",
        ]
        write_data_row(var_ws, ri, row)
        # Colour pass/fail cells
        for col_idx, pass_val in [(10, row[9]), (12, row[11]), (14, row[13])]:
            color_cell(var_ws, ri, col_idx,
                       fg=C_GREEN if "Pass" in str(pass_val) else C_RED,
                       bold=True)

    # Daily breach data below summary
    breach_start = 5 + len(loaded_syms) + 2
    var_ws.cell(row=breach_start, column=1, value="DAILY RETURNS & BREACHES").font = \
        hdr_font(color=C_ACCENT, sz=10)
    var_ws.cell(row=breach_start, column=1).fill = hdr_fill(C_BG)
    breach_start += 1

    for sym in loaded_syms:
        vr = var_results.get(sym, {})
        if not vr or "error" in vr or "returns" not in vr:
            continue
        var_ws.cell(row=breach_start, column=1, value=sym).font = hdr_font(color=C_MUTED)
        breach_start += 1
        apply_header_row(var_ws, breach_start,
                         ["Date", "Log Return", f"VaR {var_conf*100:.0f}%", "Breach"])
        breach_start += 1
        ret_s = vr["returns"]
        var_s = vr["var_series"]
        bre_s = vr["breaches"]
        common = ret_s.index.intersection(var_s.index).intersection(bre_s.index)
        for dt in common[-60:]:   # last 60 rows only to keep file size sensible
            is_breach = int(bre_s.loc[dt])
            row = [str(dt.date()),
                   round(float(ret_s.loc[dt]), 6),
                   round(float(var_s.loc[dt]), 6),
                   is_breach]
            write_data_row(var_ws, breach_start, row)
            if is_breach:
                for ci in range(1, 5):
                    var_ws.cell(row=breach_start, column=ci).font = Font(
                        name="Arial", color=C_RED, size=10)
            breach_start += 1
        breach_start += 1

    set_col_widths(var_ws, [12,12,11,10,12,11,11,12,10,10,14,10,10,12])
    var_ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET — FUTURE FORECAST
    # ══════════════════════════════════════════════════════════════════════════
    fc_ws = wb.create_sheet("Future Forecast")
    fc_ws.sheet_properties.tabColor = C_GREEN
    title_block(fc_ws, f"Future Forecasts — Next {future_days} Business Days",
                "Point forecasts with prediction bands and % change vs last price")

    fc_row = 4
    for sym in loaded_syms:
        ff = future_forecasts.get(sym, {})
        if not ff:
            continue
        fc_ws.cell(row=fc_row, column=1, value=sym).font = hdr_font(color=C_ACCENT, sz=11)
        fc_ws.cell(row=fc_row, column=1).fill = hdr_fill(C_BG)
        fc_ws.merge_cells(start_row=fc_row, start_column=1,
                          end_row=fc_row, end_column=6)
        fc_row += 1

        last_price = splits[sym]["full"].iloc[-1]
        model_used = ff.get("model", "—")
        fc_ws.cell(row=fc_row, column=1,
                   value=f"Model: {model_used} | Last Price: {last_price:.2f}").font = \
            Font(name="Arial", color=C_MUTED, size=9)
        fc_row += 1

        apply_header_row(fc_ws, fc_row,
                         ["Date", "Forecast", "Lower Band", "Upper Band",
                          "Δ vs Last", "Δ vs Last %"])
        fc_row += 1

        fp = ff["pred"]
        fl = ff.get("lower")
        fu = ff.get("upper")
        for dt, fv in fp.items():
            lo = float(fl.loc[dt]) if fl is not None and dt in fl.index else np.nan
            hi = float(fu.loc[dt]) if fu is not None and dt in fu.index else np.nan
            delta     = float(fv) - last_price
            delta_pct = delta / last_price * 100
            row = [
                str(dt.date()),
                round(float(fv), 2),
                round(lo, 2) if not np.isnan(lo) else "—",
                round(hi, 2) if not np.isnan(hi) else "—",
                round(delta, 2),
                round(delta_pct, 2),
            ]
            write_data_row(fc_ws, fc_row, row)
            # Colour direction
            fg = C_GREEN if delta_pct >= 0 else C_RED
            for ci in [5, 6]:
                color_cell(fc_ws, fc_row, ci, fg=fg, bold=True)
            fc_row += 1
        fc_row += 2

    set_col_widths(fc_ws, [14, 12, 12, 12, 12, 12])
    fc_ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SAVE TO BYTES
    # ══════════════════════════════════════════════════════════════════════════
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Report download button ────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("""
<div style="background:#161B22;border:1px solid #30363D;border-radius:10px;
            padding:20px 28px;display:flex;align-items:center;gap:16px;
            margin:8px 0 24px 0;">
    <div style="font-size:1.8rem;">📥</div>
    <div>
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.95rem;
                    font-weight:600;color:#E6EDF3;">Download Full Analysis Report</div>
        <div style="font-size:0.80rem;color:#8B949E;margin-top:3px;">
            Excel workbook · Summary · Model Metrics · Walk-Forward CV · 
            VaR Backtest · Future Forecasts
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

with st.spinner("Building Excel report…"):
    try:
        report_bytes = build_report_excel(
            loaded_syms, splits, all_results, test_results,
            final_winners, wf_results, var_results, future_forecasts,
            selected_col, future_days, var_best_method, var_conf,
        )
        fname = (f"QuantLens_Report_{'_'.join(loaded_syms)}_"
                 f"{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx")
        st.download_button(
            label="⬇️  Download Full Report (.xlsx)",
            data=report_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as _rpt_e:
        st.warning(f"Report generation failed: {_rpt_e}")


st.markdown("""
<div style="margin-top:48px; padding:16px 0; border-top:1px solid #30363D;
            text-align:center; font-family:'IBM Plex Mono',monospace;
            font-size:0.68rem; color:#8B949E; letter-spacing:0.10em;">
    QuantLens · Multi-Model Stock Forecasting System · 
    ARIMA / Log-ARIMA / Holt-Winters / XGBoost / GBM / RandomForest / LSTM / GRU · 
    Walk-Forward CV · Kupiec + Christoffersen VaR Tests · 
    For research & educational purposes only.
</div>
""", unsafe_allow_html=True)